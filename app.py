import hashlib, hmac, io, json, os, shutil, sqlite3, tempfile, uuid
from pathlib import Path
from contextlib import closing
from datetime import date, datetime
from functools import wraps
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge
from flask import Flask, abort, redirect, render_template_string, request, session, url_for, send_from_directory, send_file

APP_VERSION='7.3.2'
RELEASE_NAME='7.3-stability-and-operations'
ROOT=os.path.dirname(os.path.abspath(__file__)); DATA_DIR=os.path.join(ROOT,'data'); os.makedirs(DATA_DIR,exist_ok=True)
DB=os.environ.get('DATABASE_PATH',os.path.join(DATA_DIR,'mhoms.db'))
UPLOAD_DIR=os.environ.get('UPLOAD_PATH',os.path.join(ROOT,'uploads')); os.makedirs(UPLOAD_DIR,exist_ok=True)
app=Flask(__name__); app.secret_key=os.environ.get('SECRET_KEY','local-development-secret-change-me')
app.config.update(SESSION_COOKIE_HTTPONLY=True,SESSION_COOKIE_SAMESITE='Lax',SESSION_COOKIE_SECURE=os.environ.get('RENDER','').lower()=='true',MAX_CONTENT_LENGTH=int(os.environ.get('MAX_UPLOAD_MB','25'))*1024*1024)
ROLE_AR={'super_admin':'مدير النظام الشامل','services_manager':'مدير الخدمات المساندة','housing_manager':'مدير السكن','housing_supervisor':'مشرف السكن','housing_monitor':'مراقب السكن','maintenance_manager':'مدير الصيانة','maintenance_supervisor':'مشرف الصيانة','data_entry':'مدخل بيانات'}
REQ_AR={'transfer':'نقل عامل','final_exit':'خروج نهائي','outside_temp':'سكن خارجي مؤقت','outside_perm':'سكن خارجي دائم','add_worker':'إضافة عامل جديد','delete_worker':'حذف/أرشفة عامل'}
STATUS_AR={'pending':'بانتظار الاعتماد','approved':'معتمد','rejected':'مرفوض','new':'جديد','in_progress':'قيد التنفيذ','completed':'مكتمل','verified':'تم التحقق','closed':'مغلق','pending_maintenance':'بانتظار قبول الصيانة','accepted':'مقبول من الصيانة','awaiting_reporter':'بانتظار اعتماد مقدم البلاغ','returned':'معاد للصيانة'}
ROOM_USAGE_AR={'residential':'سكن عمال','warehouse':'مستودع','security':'حراسات الأمن الداخلي','contractor':'مقاول','administration':'إدارة','maintenance':'صيانة','laundry':'مغسلة','closed':'مغلق','out_of_service':'خارج الخدمة','other':'أخرى'}

def conn():
 c=sqlite3.connect(DB,timeout=30); c.row_factory=sqlite3.Row; c.execute('PRAGMA foreign_keys=ON'); return c

def now(): return datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
def save_upload(file_storage, prefix='ticket'):
 if not file_storage or not file_storage.filename:return None
 safe=secure_filename(file_storage.filename)
 ext=os.path.splitext(safe)[1].lower()
 allowed={'.jpg','.jpeg','.png','.webp'}
 if ext not in allowed:raise ValueError('صيغة الصورة غير مدعومة. استخدم JPG أو PNG أو WEBP.')
 os.makedirs(UPLOAD_DIR,exist_ok=True)
 name=f"{prefix}_{uuid.uuid4().hex}{ext}"
 target=os.path.join(UPLOAD_DIR,name)
 try:
  file_storage.save(target)
  if not os.path.isfile(target) or os.path.getsize(target)==0:raise OSError('تم استلام ملف فارغ')
  return name
 except Exception as exc:
  try:
   if os.path.exists(target):os.remove(target)
  except OSError:pass
  raise ValueError('تعذر حفظ الصورة على الخادم. تحقق من حجم الصورة ثم أعد المحاولة.') from exc

def cleanup_uploads(names):
 for name in names:
  if not name:continue
  try:os.remove(os.path.join(UPLOAD_DIR,os.path.basename(name)))
  except OSError:pass
def make_hash(p):
 s=os.urandom(16); d=hashlib.pbkdf2_hmac('sha256',p.encode(),s,150000).hex(); return f'{s.hex()}:{d}'
def verify(stored,p):
 try:
  s,d=stored.split(':',1); v=hashlib.pbkdf2_hmac('sha256',p.encode(),bytes.fromhex(s),150000).hex(); return hmac.compare_digest(v,d)
 except Exception:return False

def column_names(c,table): return {r[1] for r in c.execute(f'PRAGMA table_info({table})')}
def ensure_schema():
 with closing(conn()) as c:
  # additive-only migration: never drops or recreates phase-1 tables
  c.executescript('''
  CREATE TABLE IF NOT EXISTS requests(id INTEGER PRIMARY KEY AUTOINCREMENT,request_no TEXT UNIQUE,request_type TEXT,worker_id INTEGER,payload_json TEXT,requested_by INTEGER,approver_id INTEGER,status TEXT DEFAULT 'pending',decision_reason TEXT,created_at TEXT,decided_at TEXT);
  CREATE TABLE IF NOT EXISTS inspections(id INTEGER PRIMARY KEY AUTOINCREMENT,inspection_type TEXT NOT NULL,location_id TEXT NOT NULL,zone_name TEXT,inspector_id INTEGER NOT NULL,registered_count INTEGER,actual_count INTEGER,cleanliness TEXT,notes TEXT,status TEXT DEFAULT 'completed',created_at TEXT NOT NULL);
  CREATE TABLE IF NOT EXISTS inspection_people(id INTEGER PRIMARY KEY AUTOINCREMENT,inspection_id INTEGER NOT NULL,employee_no TEXT,worker_id INTEGER,registered_room TEXT,discrepancy_type TEXT);
  CREATE TABLE IF NOT EXISTS bathrooms(id INTEGER PRIMARY KEY AUTOINCREMENT,bathroom_no TEXT NOT NULL,zone_name TEXT,active INTEGER NOT NULL DEFAULT 1);
  CREATE TABLE IF NOT EXISTS maintenance_tickets(id INTEGER PRIMARY KEY AUTOINCREMENT,ticket_no TEXT NOT NULL UNIQUE,location_type TEXT NOT NULL,location_id TEXT NOT NULL,zone_name TEXT,category TEXT,description TEXT,priority TEXT DEFAULT 'normal',status TEXT DEFAULT 'new',reported_by INTEGER NOT NULL,verification_by INTEGER,assigned_to INTEGER,technician_name TEXT,part_name TEXT,completion_notes TEXT,before_photo TEXT,after_photo TEXT,created_at TEXT NOT NULL,started_at TEXT,completed_at TEXT,verified_at TEXT,closed_at TEXT);
  CREATE TABLE IF NOT EXISTS ticket_updates(id INTEGER PRIMARY KEY AUTOINCREMENT,ticket_id INTEGER NOT NULL,user_id INTEGER NOT NULL,action TEXT NOT NULL,notes TEXT,photo_path TEXT,created_at TEXT NOT NULL);
  CREATE TABLE IF NOT EXISTS audit_logs(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER,username TEXT,action TEXT,entity_type TEXT,entity_id INTEGER,details_json TEXT,created_at TEXT);
  CREATE TABLE IF NOT EXISTS room_usage_history(id INTEGER PRIMARY KEY AUTOINCREMENT,room_id INTEGER NOT NULL,old_usage TEXT,new_usage TEXT NOT NULL,reason TEXT,changed_by INTEGER,changed_at TEXT NOT NULL);
  CREATE TABLE IF NOT EXISTS import_batches(id INTEGER PRIMARY KEY AUTOINCREMENT,file_name TEXT,rows_total INTEGER,workers_imported INTEGER,special_rooms INTEGER,created_by INTEGER,created_at TEXT);
  CREATE TABLE IF NOT EXISTS bathroom_reports(id INTEGER PRIMARY KEY AUTOINCREMENT,report_no TEXT NOT NULL UNIQUE,bathroom_no TEXT NOT NULL,zone_name TEXT,issue_type TEXT NOT NULL,description TEXT,priority TEXT DEFAULT 'normal',status TEXT DEFAULT 'new',reported_by INTEGER NOT NULL,maintenance_ticket_id INTEGER,created_at TEXT NOT NULL,closed_at TEXT);
  CREATE TABLE IF NOT EXISTS worker_change_requests(id INTEGER PRIMARY KEY AUTOINCREMENT,request_no TEXT NOT NULL UNIQUE,change_type TEXT NOT NULL,worker_id INTEGER,employee_no TEXT,iqama_no TEXT,full_name TEXT,nationality TEXT,profession TEXT,zone TEXT,room_no TEXT,reason TEXT,requested_by INTEGER NOT NULL,status TEXT DEFAULT 'pending',decided_by INTEGER,decision_reason TEXT,created_at TEXT NOT NULL,decided_at TEXT);
  CREATE TABLE IF NOT EXISTS ticket_photos(id INTEGER PRIMARY KEY AUTOINCREMENT,ticket_id INTEGER NOT NULL,photo_path TEXT NOT NULL,photo_kind TEXT DEFAULT 'report',uploaded_by INTEGER NOT NULL,created_at TEXT NOT NULL);
  CREATE TABLE IF NOT EXISTS housing_kit_deliveries(id INTEGER PRIMARY KEY AUTOINCREMENT,worker_id INTEGER NOT NULL,bed INTEGER DEFAULT 0,mattress INTEGER DEFAULT 0,pillow INTEGER DEFAULT 0,sheet INTEGER DEFAULT 0,blanket INTEGER DEFAULT 0,delivered_by INTEGER NOT NULL,notes TEXT,delivered_at TEXT NOT NULL,updated_at TEXT);
  CREATE TABLE IF NOT EXISTS housing_actions(id INTEGER PRIMARY KEY AUTOINCREMENT,action_no TEXT UNIQUE,action_type TEXT NOT NULL,worker_id INTEGER,employee_no TEXT,iqama_no TEXT,full_name TEXT,nationality TEXT,profession TEXT,phone TEXT,source_room TEXT,target_room TEXT,reason TEXT,kit_bed INTEGER DEFAULT 0,kit_mattress INTEGER DEFAULT 0,kit_pillow INTEGER DEFAULT 0,kit_sheet INTEGER DEFAULT 0,kit_blanket INTEGER DEFAULT 0,requested_by INTEGER NOT NULL,source_supervisor_id INTEGER,source_supervisor_status TEXT,target_supervisor_id INTEGER,target_supervisor_status TEXT,final_status TEXT DEFAULT 'pending_supervisor',final_decided_by INTEGER,decision_reason TEXT,created_at TEXT NOT NULL,source_decided_at TEXT,target_decided_at TEXT,final_decided_at TEXT);
  CREATE INDEX IF NOT EXISTS idx_housing_actions_status ON housing_actions(final_status,requested_by);
  CREATE INDEX IF NOT EXISTS idx_ticket_photos_ticket ON ticket_photos(ticket_id);
  CREATE INDEX IF NOT EXISTS idx_kit_worker ON housing_kit_deliveries(worker_id);
  CREATE INDEX IF NOT EXISTS idx_requests_status ON requests(status);
  CREATE INDEX IF NOT EXISTS idx_inspections_location ON inspections(inspection_type,location_id,created_at);
  CREATE INDEX IF NOT EXISTS idx_tickets_status ON maintenance_tickets(status);
  CREATE INDEX IF NOT EXISTS idx_bathroom_reports_reporter ON bathroom_reports(reported_by,status);
  CREATE INDEX IF NOT EXISTS idx_worker_change_status ON worker_change_requests(status,requested_by);
  CREATE TABLE IF NOT EXISTS bathroom_complexes(id INTEGER PRIMARY KEY AUTOINCREMENT,complex_no TEXT NOT NULL UNIQUE,zone_name TEXT NOT NULL,name TEXT,active INTEGER DEFAULT 1);
  CREATE TABLE IF NOT EXISTS bathroom_assets(id INTEGER PRIMARY KEY AUTOINCREMENT,complex_id INTEGER NOT NULL,asset_type TEXT NOT NULL,asset_no INTEGER NOT NULL,supervisor_id INTEGER,status TEXT DEFAULT 'active',notes TEXT,UNIQUE(complex_id,asset_type,asset_no));
  CREATE TABLE IF NOT EXISTS bathroom_asset_inspections(id INTEGER PRIMARY KEY AUTOINCREMENT,asset_id INTEGER NOT NULL,inspector_id INTEGER NOT NULL,condition_status TEXT,notes TEXT,photo_path TEXT,maintenance_required INTEGER DEFAULT 0,maintenance_ticket_id INTEGER,created_at TEXT NOT NULL);
  CREATE TABLE IF NOT EXISTS absence_reports(id INTEGER PRIMARY KEY AUTOINCREMENT,worker_id INTEGER NOT NULL,room_no TEXT NOT NULL,reported_by INTEGER NOT NULL,absence_duration TEXT,notes TEXT,status TEXT DEFAULT 'new',manager_id INTEGER,manager_notes TEXT,created_at TEXT NOT NULL,updated_at TEXT);
  CREATE TABLE IF NOT EXISTS room_events(id INTEGER PRIMARY KEY AUTOINCREMENT,room_no TEXT NOT NULL,event_type TEXT NOT NULL,description TEXT,user_id INTEGER,related_id INTEGER,created_at TEXT NOT NULL);
  CREATE TABLE IF NOT EXISTS worker_events(id INTEGER PRIMARY KEY AUTOINCREMENT,worker_id INTEGER NOT NULL,event_type TEXT NOT NULL,description TEXT,user_id INTEGER,related_id INTEGER,created_at TEXT NOT NULL);
  CREATE INDEX IF NOT EXISTS idx_bath_assets_supervisor ON bathroom_assets(supervisor_id,asset_type);
  CREATE INDEX IF NOT EXISTS idx_absence_status ON absence_reports(status,reported_by);
  CREATE INDEX IF NOT EXISTS idx_room_events_room ON room_events(room_no,created_at);
  CREATE INDEX IF NOT EXISTS idx_worker_events_worker ON worker_events(worker_id,created_at);
  CREATE TABLE IF NOT EXISTS project_workers(employee_no TEXT PRIMARY KEY,full_name TEXT,profession TEXT,shift TEXT,rest_day TEXT,source_sheet TEXT,source_location TEXT,updated_at TEXT NOT NULL);
  CREATE TABLE IF NOT EXISTS attendance_batches(id INTEGER PRIMARY KEY AUTOINCREMENT,batch_no TEXT NOT NULL UNIQUE,absence_date TEXT NOT NULL,shift_label TEXT,created_by INTEGER NOT NULL,status TEXT DEFAULT 'draft',source_files_json TEXT,created_at TEXT NOT NULL,completed_at TEXT);
  CREATE TABLE IF NOT EXISTS attendance_entries(id INTEGER PRIMARY KEY AUTOINCREMENT,batch_id INTEGER NOT NULL,employee_no TEXT NOT NULL,full_name TEXT,profession TEXT,shift TEXT,rest_day TEXT,is_resident INTEGER DEFAULT 0,room_no TEXT,zone TEXT,source_files TEXT,previous_absences INTEGER DEFAULT 0,reason TEXT,notes TEXT,updated_at TEXT,UNIQUE(batch_id,employee_no));
  CREATE INDEX IF NOT EXISTS idx_attendance_batch_status ON attendance_batches(status,created_by,absence_date);
  CREATE INDEX IF NOT EXISTS idx_attendance_employee ON attendance_entries(employee_no,batch_id);
  ''')
  # future-safe additive columns
  for table,defs in {
   'requests':{'supervisor_id':'INTEGER','supervisor_decision_at':'TEXT'},
   'inspections':{'week_key':'TEXT','maintenance_ticket_id':'INTEGER'},
   'rooms':{'usage_type':"TEXT DEFAULT 'residential'",'length_m':'REAL','width_m':'REAL','area_m2':'REAL','status':"TEXT DEFAULT 'active'",'notes':'TEXT','updated_at':'TEXT','supervisor_id':'INTEGER','sector_name':'TEXT'},
   'workers':{'archived':'INTEGER DEFAULT 0','status':"TEXT DEFAULT 'active'",'source_row':'INTEGER','import_batch_id':'INTEGER','phone':'TEXT','updated_at':'TEXT'},
   'worker_change_requests':{'phone':'TEXT','kit_bed':'INTEGER DEFAULT 0','kit_mattress':'INTEGER DEFAULT 0','kit_pillow':'INTEGER DEFAULT 0','kit_sheet':'INTEGER DEFAULT 0','kit_blanket':'INTEGER DEFAULT 0'}
  }.items():
   cols=column_names(c,table)
   for name,typ in defs.items():
    if name not in cols:c.execute(f'ALTER TABLE {table} ADD COLUMN {name} {typ}')
  # Final phase-4 bootstrap: geographic/even room ownership, protected management rooms, and numbered bathroom assets.
  supervisors=[r['id'] for r in c.execute("SELECT id FROM users WHERE role='housing_supervisor' AND active=1 ORDER BY id").fetchall()]
  rooms_for_assignment=c.execute("SELECT id,room_no,zone FROM rooms WHERE NOT (CAST(room_no AS INTEGER) BETWEEN 3801 AND 3816) ORDER BY CAST(zone AS INTEGER),CAST(room_no AS INTEGER),room_no").fetchall()
  if supervisors and rooms_for_assignment:
   base=len(rooms_for_assignment)//len(supervisors); extra=len(rooms_for_assignment)%len(supervisors); pos=0
   for idx,sid in enumerate(supervisors):
    size=base+(1 if idx<extra else 0); batch=rooms_for_assignment[pos:pos+size]; pos+=size
    for rr in batch:c.execute("UPDATE rooms SET supervisor_id=?,sector_name=?,updated_at=COALESCE(updated_at,?) WHERE id=?",(sid,f'قطاع {idx+1}',now(),rr['id']))
   c.execute("UPDATE rooms SET supervisor_id=NULL,sector_name='إدارة السكن' WHERE CAST(room_no AS INTEGER) BETWEEN 3801 AND 3816")
  for z in ('1','2','3','4'):
   c.execute("INSERT OR IGNORE INTO bathroom_complexes(complex_no,zone_name,name) VALUES(?,?,?)",(f'BC-{z}',z,f'مجمع دورات المياه - زون {z}'))
  complexes=c.execute("SELECT * FROM bathroom_complexes ORDER BY id").fetchall()
  for comp in complexes:
   zone_sups=[r['id'] for r in c.execute("SELECT DISTINCT supervisor_id id FROM rooms WHERE CAST(zone AS TEXT)=? AND supervisor_id IS NOT NULL ORDER BY supervisor_id",(comp['zone_name'],)).fetchall()] or supervisors
   for typ,total in (('toilet',156),('basin',80)):
    for n in range(1,total+1):
     sid=zone_sups[(n-1)%len(zone_sups)] if zone_sups else None
     c.execute("INSERT OR IGNORE INTO bathroom_assets(complex_id,asset_type,asset_no,supervisor_id) VALUES(?,?,?,?)",(comp['id'],typ,n,sid))
  # Built-in full-access test account. Password can be overridden at deployment.
  admin_password=os.environ.get('SUPER_ADMIN_PASSWORD','Admin@73')
  existing=c.execute("SELECT id FROM users WHERE employee_no='admin' OR username='admin'").fetchone()
  if existing:
   c.execute("UPDATE users SET role='super_admin',active=1 WHERE id=?",(existing['id'],))
  else:
   c.execute("INSERT INTO users(employee_no,username,display_name,password_hash,role,preferred_lang,active,must_change_password) VALUES(?,?,?,?,?,?,?,?)",('admin','admin','مدير النظام الشامل',make_hash(admin_password),'super_admin','ar',1,1))
  c.commit()
ensure_schema()

def current_user():
 uid=session.get('uid')
 if not uid:return None
 with closing(conn()) as c:return c.execute('SELECT * FROM users WHERE id=? AND active=1',(uid,)).fetchone()
def login_required(fn):
 @wraps(fn)
 def inner(*a,**k):
  u=current_user()
  if not u:return redirect(url_for('login'))
  if u['must_change_password'] and request.endpoint not in {'change_password','logout','static'}:return redirect(url_for('change_password'))
  return fn(*a,**k)
 return inner
def is_admin(u):return u['role'] in ('super_admin','services_manager','housing_manager')
def can_export(u):return u['role'] in ('super_admin','services_manager','housing_manager')
def can_maintenance(u):return u['role'] in ('super_admin','services_manager','housing_manager','maintenance_manager','maintenance_supervisor')
def can_create_housing(u):return u['role'] in ('housing_supervisor','housing_monitor','data_entry')
def can_attendance(u):return u['role'] in ('super_admin','services_manager','housing_manager','housing_supervisor','housing_monitor')
def is_maintenance_only(u):return u['role'] in ('maintenance_manager','maintenance_supervisor')
def assigned_clause(u,alias='r'):
 if u['role']!='housing_supervisor':return '1=1',[]
 return f'{alias}.supervisor_id=?',[u['id']]

def assigned_zones(c,u):
 if u['role']!='housing_supervisor':
  return [str(r['zone_name']) for r in c.execute("SELECT DISTINCT zone_name FROM bathroom_complexes WHERE active=1 ORDER BY CAST(zone_name AS INTEGER),zone_name").fetchall()]
 return [str(r['zone']) for r in c.execute("SELECT DISTINCT zone FROM rooms WHERE supervisor_id=? AND zone IS NOT NULL ORDER BY CAST(zone AS INTEGER),zone",(u['id'],)).fetchall()]

def supervisor_for_room(room_no):
 try:n=int(room_no)
 except:return None
 with closing(conn()) as c:return c.execute('''SELECT u.* FROM users u JOIN assignments a ON a.user_id=u.id WHERE u.active=1 AND u.role='housing_supervisor' AND ? BETWEEN MIN(a.room_start,a.room_end) AND MAX(a.room_start,a.room_end) ORDER BY a.id LIMIT 1''',(n,)).fetchone()
def audit(c,u,action,etype,eid,details=None):
 c.execute('INSERT INTO audit_logs(user_id,username,action,entity_type,entity_id,details_json,created_at) VALUES(?,?,?,?,?,?,?)',(u['id'] if u else None,u['employee_no'] if u else None,action,etype,eid,json.dumps(details or {},ensure_ascii=False),now()))


def worker_export_rows(c, q='', zone='', status=''):
 sql="""SELECT w.employee_no,w.iqama_no,w.full_name,w.nationality,w.profession,w.phone,
               w.zone,w.room_no,w.status,w.created_at,w.updated_at,
               r.sector_name,r.capacity,r.usage_type,r.status AS room_status,
               u.display_name AS supervisor_name
        FROM workers w
        LEFT JOIN rooms r ON r.room_no=w.room_no
        LEFT JOIN users u ON u.id=r.supervisor_id
        WHERE COALESCE(w.archived,0)=0"""
 params=[]
 if q:
  sql+=' AND (w.employee_no LIKE ? OR w.iqama_no LIKE ? OR w.full_name LIKE ? OR w.room_no LIKE ? OR w.nationality LIKE ? OR w.profession LIKE ?)'
  like=f'%{q}%';params += [like]*6
 if zone:
  sql+=' AND CAST(w.zone AS TEXT)=?';params.append(str(zone))
 if status:
  sql+=" AND COALESCE(w.status,'active')=?";params.append(status)
 sql+=' ORDER BY CAST(w.zone AS INTEGER),CAST(w.room_no AS INTEGER),w.full_name'
 return c.execute(sql,params).fetchall()

def ar_text(value):
 value='' if value is None else str(value)
 try:
  import arabic_reshaper
  reshaped=arabic_reshaper.reshape(value)
  try:
   from bidi.algorithm import get_display
   return get_display(reshaped)
  except Exception:
   return reshaped[::-1]
 except Exception:
  return value

def export_filters():
 return request.args.get('q','').strip(),request.args.get('zone','').strip(),request.args.get('status','').strip()


I18N={
 'ar':{'home':'الرئيسية','rooms':'الغرف','workers':'العمال','inspections':'الجولات','maintenance':'الصيانة','requests':'الطلبات','occupancy':'إدارة الإشغال','map':'خريطة السكن','notifications':'الإشعارات','logout':'خروج','login':'تسجيل الدخول','employee_no':'الرقم الوظيفي','password':'كلمة المرور','enter':'دخول','choose_language':'اختر اللغة','welcome':'مرحبًا بك في نظام MAG CAMP','system_name':'نظام إدارة سكن ولي العهد','change_password':'تغيير كلمة المرور','worker_changes':'التسكين الجديد / حذف عامل','users':'المستخدمون','audit':'سجل العمليات','maintenance_dashboard':'لوحة الصيانة','search':'البحث الشامل'},
 'en':{'home':'Dashboard','rooms':'Rooms','workers':'Workers','inspections':'Inspections','maintenance':'Maintenance','requests':'Requests','occupancy':'Occupancy Management','map':'Occupancy Map','notifications':'Notifications','logout':'Logout','login':'Sign In','employee_no':'Employee Number','password':'Password','enter':'Sign In','choose_language':'Select Language','welcome':'Welcome to MAG CAMP','system_name':'Wali Al Ahd Camp Management System','change_password':'Change Password','worker_changes':'New Housing / Remove Worker','users':'Users','audit':'Audit Log','maintenance_dashboard':'Maintenance Dashboard','search':'Global Search'}
}
def lang(): return session.get('lang') if session.get('lang') in ('ar','en') else None
def tr(key): return I18N.get(lang() or 'ar',I18N['ar']).get(key,key)

BASE='''<!doctype html><html lang="{{lang_code}}" dir="{{direction}}"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1"><title>{{title}}</title><style>
:root{--green:#123b32;--green2:#1e5a4b;--bg:#f4f6f8;--red:#a52222;--amber:#a66a00}
*{box-sizing:border-box}html,body{max-width:100%;overflow-x:hidden}body{font-family:Tahoma,Arial;background:var(--bg);margin:0;color:#24302d}h1,h2,h3{overflow-wrap:anywhere;word-break:break-word}.top{background:var(--green);color:#fff;padding:12px 18px;display:flex;justify-content:space-between;align-items:center;gap:12px;position:sticky;top:0;z-index:20}.brand{display:flex;align-items:center;gap:10px}.brand img{width:74px;height:52px;object-fit:contain;background:#fff;border-radius:9px;padding:5px}.donuts{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:14px}.donut-card{background:#fff;border-radius:14px;padding:16px;text-align:center;box-shadow:0 2px 8px #d8dddd}.donut{--p:0;--c:#25845f;width:112px;height:112px;border-radius:50%;margin:8px auto;background:conic-gradient(var(--c) calc(var(--p)*1%),#e7ecea 0);display:grid;place-items:center}.donut:after{content:attr(data-label);width:76px;height:76px;border-radius:50%;background:white;display:grid;place-items:center;font-weight:bold;font-size:18px}.kanban{display:grid;grid-template-columns:repeat(5,minmax(220px,1fr));gap:12px;overflow:auto}.kanban-col{background:#eef2f1;border-radius:12px;padding:10px;min-height:250px}.ticket-card{background:#fff;border-radius:10px;padding:11px;margin-bottom:9px;box-shadow:0 1px 4px #ccd}.photo-stage{display:grid;grid-template-columns:1fr;gap:8px}.photo-stage img{width:100%;max-height:360px;object-fit:contain;background:#f2f2f2;border-radius:12px}.wrap{display:flex;min-height:calc(100vh - 70px)}aside{width:235px;background:#fff;padding:14px;box-shadow:0 0 8px #ccc;flex-shrink:0}aside a{display:block;padding:11px;color:var(--green);text-decoration:none;border-bottom:1px solid #eee;border-radius:7px}aside a:hover{background:#edf4f1}main{flex:1;padding:22px;min-width:0}.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(165px,1fr));gap:12px}.card{background:#fff;border-radius:13px;padding:17px;box-shadow:0 2px 8px #d8dddd;margin-bottom:14px}.num{font-size:27px;font-weight:bold;margin-top:5px}.muted{color:#68736f;font-size:13px}.tbl-wrap{overflow:auto;background:#fff;border-radius:12px}.tbl{width:100%;border-collapse:collapse;min-width:720px}.tbl th,.tbl td{padding:10px;border-bottom:1px solid #eee;text-align:right;font-size:14px}.tbl th{background:#edf4f1;position:sticky;top:0}.search,input,select,textarea{padding:11px;border:1px solid #ccd4d1;border-radius:8px;max-width:100%;font:inherit}.search{width:min(420px,100%);margin-bottom:10px}.btn{display:inline-block;background:var(--green);color:#fff;border:0;border-radius:8px;padding:10px 16px;text-decoration:none;cursor:pointer}.btn2{background:#6a7c76}.danger{background:var(--red)}.login{max-width:390px;margin:9vh auto;background:#fff;padding:28px;border-radius:14px;box-shadow:0 3px 18px #bbb}.err{color:#b00020}.ok{color:#0a6b3c}.badge{padding:4px 8px;border-radius:10px;background:#e5eee9;white-space:nowrap}.badge.red{background:#fde7e7;color:#8d1616}.badge.amber{background:#fff1d6;color:#805000}.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px}.field label{display:block;margin:5px 0}.field input,.field select,.field textarea{width:100%}.mobile-nav{display:none}.room-map{display:grid;grid-template-columns:repeat(auto-fill,minmax(115px,1fr));gap:10px}.room-tile{padding:14px;border-radius:12px;color:#fff;text-decoration:none;min-height:88px;display:flex;flex-direction:column;justify-content:space-between}.room-tile.green{background:#25845f}.room-tile.yellow{background:#d49a18}.room-tile.red{background:#b83232}.room-tile.black{background:#222}.room-tile.blue{background:#2b6fb3}.timeline{border-right:3px solid #1e5a4b;padding-right:18px}.timeline-item{background:#fff;padding:12px;margin:0 0 12px;border-radius:10px;box-shadow:0 1px 5px #ddd}
@media(max-width:800px){
.top{align-items:center;padding:9px 10px;min-height:58px}.top .userline{font-size:11px;text-align:left;line-height:1.35}.brand{min-width:0}.brand div{min-width:0}.brand b{font-size:14px}.brand small{display:none}.brand img{width:38px;height:38px;padding:3px}.wrap{display:block;min-height:0}.desktop-nav{display:none}
.mobile-nav{display:flex;align-items:center;gap:7px;background:#fff;padding:7px 8px;position:sticky;top:58px;z-index:15;box-shadow:0 2px 7px #ddd;overflow-x:auto;overflow-y:hidden;white-space:nowrap;-webkit-overflow-scrolling:touch;scrollbar-width:none;min-height:49px}.mobile-nav::-webkit-scrollbar{display:none}.attendance-title{font-size:24px;line-height:1.35}.attendance-title small{font-size:15px;direction:ltr;display:inline-block;max-width:100%;overflow-wrap:anywhere}.mobile-nav a{flex:0 0 auto;text-align:center;text-decoration:none;color:var(--green);font-size:12px;line-height:1;padding:10px 12px;border-radius:8px;background:#f1f5f3}
main{padding:12px 10px;margin:0;min-height:0;width:100%;max-width:100%;overflow:hidden}.tbl-wrap{width:100%;max-width:100%;overflow-x:auto;-webkit-overflow-scrolling:touch}.tbl{min-width:980px}.cards{grid-template-columns:repeat(2,minmax(0,1fr));gap:9px}.card{padding:13px;margin-bottom:10px}.num{font-size:23px}.tbl{min-width:650px}}
@media(max-width:380px){.cards{grid-template-columns:1fr}.mobile-nav a{padding:9px 10px}.top .userline{max-width:135px}}
</style></head><body><div class="top"><div class="brand"><img src="{{url_for('static',filename='mag_logo.png')}}" alt="MAG"><div><b>MAG CAMP</b><br><small>{{t.system_name}} — Enterprise 7.3</small></div></div>{% if u %}<div class="userline">{{u['display_name']}}<br><small>{{roles.get(u['role'],u['role'])}} | <a style="color:white" href="{{url_for('logout')}}">{{t.logout}}</a></small></div>{% endif %}</div>{% if u %}<nav class="mobile-nav"><a href="{{url_for('dashboard')}}">{{t.home}}</a><a href="{{url_for('notifications')}}">{{t.notifications}}</a>{% if not maintenance_only %}<a href="{{url_for('global_search')}}">{{t.search}}</a><a href="{{url_for('rooms')}}">{{t.rooms}}</a><a href="{{url_for('occupancy_management')}}">{{t.occupancy}}</a><a href="{{url_for('workers')}}">{{t.workers}}</a>{% if u.role=='housing_supervisor' %}<a href="{{url_for('inspections')}}">{{t.inspections}}</a>{% endif %}<a href="{{url_for('sector_dashboard')}}">القطاعات</a><a href="{{url_for('worker_change_requests')}}">إدارة الطلبات</a><a href="{{url_for('absence_reports_list')}}">بلاغات عدم التواجد</a>{% if can_attendance %}<a href="{{url_for('attendance_batches')}}">حصر الغياب</a>{% endif %}{% if admin %}<a href="{{url_for('reports_center')}}">مركز التقارير</a>{% endif %}{% endif %}<a href="{{url_for('tickets')}}">{{t.maintenance}}</a></nav><div class="wrap"><aside class="desktop-nav"><a href="{{url_for('dashboard')}}">{{t.home}}</a><a href="{{url_for('notifications')}}">{{t.notifications}}</a>{% if not maintenance_only %}<a href="{{url_for('global_search')}}">{{t.search}}</a><a href="{{url_for('workers')}}">{{t.workers}}</a><a href="{{url_for('rooms')}}">{{t.rooms}}</a><a href="{{url_for('occupancy_management')}}">{{t.occupancy}}</a>{% if u.role=='housing_supervisor' %}<a href="{{url_for('inspections')}}">{{t.inspections}}</a>{% endif %}<a href="{{url_for('sector_dashboard')}}">القطاعات</a>{% endif %}{% if not maintenance_only %}<a href="{{url_for('worker_change_requests')}}">إدارة الطلبات</a><a href="{{url_for('absence_reports_list')}}">بلاغات عدم التواجد</a>{% if can_attendance %}<a href="{{url_for('attendance_batches')}}">حصر الغياب</a>{% endif %}{% if admin %}<a href="{{url_for('reports_center')}}">مركز التقارير</a>{% endif %}{% endif %}<a href="{{url_for('tickets')}}">{{t.maintenance}}</a>{% if u.role in ('super_admin','maintenance_manager','maintenance_supervisor','housing_manager','services_manager') %}<a href="{{url_for('maintenance_dashboard')}}">{{t.maintenance_dashboard}}</a>{% endif %}{% if admin %}<a href="{{url_for('users')}}">{{t.users}}</a><a href="{{url_for('audit_logs')}}">{{t.audit}}</a><a href="{{url_for('backup_restore')}}">النسخ الاحتياطي</a>{% endif %}<a href="{{url_for('change_password')}}">{{t.change_password}}</a></aside><main>{{body|safe}}</main></div>{% else %}{{body|safe}}{% endif %}</body></html>'''
def page(body,title='MAG CAMP',user=None,**ctx):
 lc=lang() or 'ar'; t=I18N[lc]
 return render_template_string(BASE,title=title,body=render_template_string(body,t=t,lang_code=lc,**ctx),u=user,roles=ROLE_AR,admin=bool(user and is_admin(user)),maintenance_only=bool(user and is_maintenance_only(user)),can_attendance=bool(user and can_attendance(user)),t=t,lang_code=lc,direction='rtl' if lc=='ar' else 'ltr')

@app.get('/health')
def health():
 try:
  with closing(conn()) as c:
   c.execute('SELECT 1'); counts={t:c.execute(f'SELECT COUNT(*) FROM {t}').fetchone()[0] for t in ('users','rooms','workers','requests','inspections','maintenance_tickets')}
  return {'status':'ok','database':'ok','version':APP_VERSION,'release':RELEASE_NAME,'python':os.sys.version.split()[0],'counts':counts},200
 except Exception as e:return {'status':'error','database':'unavailable','message':str(e)},503
@app.route('/language',methods=['GET','POST'])
def choose_language():
 if request.method=='POST':
  choice=request.form.get('language')
  if choice in ('ar','en'):
   session['lang']=choice
   return redirect(url_for('login'))
 return render_template_string('''<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Select Language</title><style>body{font-family:Tahoma,Arial;background:#f4f6f8;margin:0}.box{max-width:520px;margin:12vh auto;background:#fff;padding:35px;border-radius:18px;box-shadow:0 5px 24px #cdd4d1;text-align:center}.logo{width:105px}.choices{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:25px}.choice{border:0;border-radius:12px;padding:18px;background:#123b32;color:white;font-size:18px;cursor:pointer}@media(max-width:540px){.box{margin:8vh 14px}.choices{grid-template-columns:1fr}}</style></head><body><div class="box"><img class="logo" src="{{url_for('static',filename='mag_logo.png')}}"><h1>MAG CAMP</h1><p>اختر اللغة | Select Language</p><form method="post" class="choices"><button class="choice" name="language" value="ar">العربية</button><button class="choice" name="language" value="en">English</button></form></div></body></html>''')

@app.route('/login',methods=['GET','POST'])
def login():
 if not lang():
  if request.method=='POST': session['lang']='ar'
  else: return redirect(url_for('choose_language'))
 if current_user():return redirect(url_for('dashboard'))
 error=''
 if request.method=='POST':
  eno=request.form.get('employee_no','').strip(); pw=request.form.get('password','')
  with closing(conn()) as c:
   u=c.execute('SELECT * FROM users WHERE employee_no=? AND active=1',(eno,)).fetchone()
   if u and verify(u['password_hash'],pw):
    chosen_lang=lang() or 'ar';session.clear();session['lang']=chosen_lang;session['uid']=u['id'];c.execute("UPDATE users SET last_login=datetime('now') WHERE id=?",(u['id'],));c.commit();return redirect(url_for('change_password') if u['must_change_password'] else url_for('dashboard'))
  error='الرقم الوظيفي أو كلمة المرور غير صحيحة' if lang()=='ar' else 'Invalid employee number or password'
 return page('''<div class="login"><div style="text-align:center"><img src="{{url_for('static',filename='mag_logo.png')}}" style="width:90px"><h2>{{t.login}}</h2></div>{% if error %}<p class="err">{{error}}</p>{% endif %}<form method="post"><input class="search" name="employee_no" placeholder="{{t.employee_no}}" required><input class="search" type="password" name="password" placeholder="{{t.password}}" required><button class="btn">{{t.enter}}</button></form><p><a href="{{url_for('choose_language')}}">{{t.choose_language}}</a></p></div>''',tr('login'),error=error)
@app.get('/logout')
def logout():
 lc=lang();session.clear();session['lang']=lc or 'ar';return redirect(url_for('login'))
@app.route('/change-password',methods=['GET','POST'])
@login_required
def change_password():
 u=current_user();err='';ok=''
 if request.method=='POST':
  cur=request.form.get('current_password','');new=request.form.get('new_password','');conf=request.form.get('confirm_password','')
  if not verify(u['password_hash'],cur):err='كلمة المرور الحالية غير صحيحة'
  elif len(new)<6:err='كلمة المرور الجديدة يجب ألا تقل عن 6 أحرف أو أرقام'
  elif new!=conf:err='تأكيد كلمة المرور غير مطابق'
  else:
   with closing(conn()) as c:c.execute('UPDATE users SET password_hash=?,must_change_password=0 WHERE id=?',(make_hash(new),u['id']));audit(c,u,'change_password','user',u['id']);c.commit();ok='تم تغيير كلمة المرور بنجاح'
 return page('''<div class="card"><h2>تغيير كلمة المرور</h2>{% if err %}<p class="err">{{err}}</p>{% endif %}{% if ok %}<p class="ok">{{ok}}</p><a class="btn" href="{{url_for('dashboard')}}">{{t.home}}</a>{% else %}<form method="post"><input class="search" type="password" name="current_password" placeholder="الحالية" required><input class="search" type="password" name="new_password" placeholder="الجديدة" required><input class="search" type="password" name="confirm_password" placeholder="التأكيد" required><button class="btn">حفظ</button></form>{% endif %}</div>''','تغيير كلمة المرور',u,err=err,ok=ok)

@app.get('/')
@login_required
def dashboard():
 u=current_user();cl,args=assigned_clause(u,'r');week=date.today().strftime('%Y-W%W')
 with closing(conn()) as c:
  room_stats=c.execute(f'''SELECT r.room_no,r.zone,r.capacity,r.usage_type,r.status,COUNT(w.id) occupied,(r.capacity-COUNT(w.id)) free_beds FROM rooms r LEFT JOIN workers w ON w.room_no=r.room_no AND w.archived=0 WHERE {cl} GROUP BY r.id ORDER BY CAST(r.room_no AS INTEGER)''',args).fetchall()
  residential=[x for x in room_stats if (x['usage_type'] or 'residential')=='residential']
  rooms=len(residential);workers=sum(x['occupied'] for x in residential);capacity=sum(x['capacity'] or 0 for x in residential)
  vacant=sum(1 for x in residential if x['occupied']==0); crowded=sum(1 for x in residential if x['occupied']>(x['capacity'] or 0)); occupied=sum(1 for x in residential if x['occupied']>0 and x['occupied']<=x['capacity'])
  closed=sum(1 for x in room_stats if (x['status'] or '') in ('closed','out_of_service') or (x['usage_type'] or '')=='closed'); maintenance=sum(1 for x in room_stats if (x['usage_type'] or '')=='maintenance')
  usage_counts={k:sum(1 for x in room_stats if (x['usage_type'] or 'residential')==k) for k in ('warehouse','contractor','security')}
  zones_for_user=assigned_zones(c,u)
  if u['role']=='housing_supervisor' and zones_for_user:
   marks=','.join('?' for _ in zones_for_user)
   bathroom_complexes_count=c.execute(f"SELECT COUNT(DISTINCT a.complex_id) FROM bathroom_assets a JOIN bathroom_complexes b ON b.id=a.complex_id WHERE CAST(b.zone_name AS TEXT) IN ({marks})",zones_for_user).fetchone()[0]
   toilets_count=c.execute(f"SELECT COUNT(*) FROM bathroom_assets a JOIN bathroom_complexes b ON b.id=a.complex_id WHERE a.asset_type='toilet' AND CAST(b.zone_name AS TEXT) IN ({marks})",zones_for_user).fetchone()[0]
   basins_count=c.execute(f"SELECT COUNT(*) FROM bathroom_assets a JOIN bathroom_complexes b ON b.id=a.complex_id WHERE a.asset_type='basin' AND CAST(b.zone_name AS TEXT) IN ({marks})",zones_for_user).fetchone()[0]
  else:
   bathroom_complexes_count=c.execute("SELECT COUNT(DISTINCT complex_id) FROM bathroom_assets").fetchone()[0]
   toilets_count=c.execute("SELECT COUNT(*) FROM bathroom_assets WHERE asset_type='toilet'").fetchone()[0]
   basins_count=c.execute("SELECT COUNT(*) FROM bathroom_assets WHERE asset_type='basin'").fetchone()[0]
  absence_open=c.execute("SELECT COUNT(*) FROM absence_reports WHERE status NOT IN ('closed','rejected')"+(" AND reported_by=?" if u['role']=='housing_supervisor' else ''),([u['id']] if u['role']=='housing_supervisor' else [])).fetchone()[0]
  tparams=[u['id']] if u['role'] in ('housing_supervisor','housing_monitor') else []
  twhere=' AND reported_by=?' if tparams else ''
  open_t=c.execute("SELECT COUNT(*) FROM maintenance_tickets WHERE status NOT IN ('closed')"+twhere,tparams).fetchone()[0]
  closed_t=c.execute("SELECT COUNT(*) FROM maintenance_tickets WHERE status='closed'"+twhere,tparams).fetchone()[0]
  done=c.execute("SELECT COUNT(DISTINCT location_id) FROM inspections WHERE inspection_type='room' AND week_key=?"+(" AND inspector_id=?" if u['role']=='housing_supervisor' else ''),([week,u['id']] if u['role']=='housing_supervisor' else [week])).fetchone()[0]
  assigned_total=len(residential); progress=round(done*100/assigned_total,1) if assigned_total else 0
  pending_housing=c.execute("SELECT COUNT(*) FROM housing_actions WHERE final_status NOT IN ('approved','rejected')").fetchone()[0] if is_admin(u) else c.execute("SELECT COUNT(*) FROM housing_actions WHERE requested_by=? AND final_status NOT IN ('approved','rejected')",(u['id'],)).fetchone()[0]
  latest_tickets=c.execute("SELECT id,ticket_no,location_id,status,created_at FROM maintenance_tickets ORDER BY id DESC LIMIT 8").fetchall()
  latest_actions=c.execute("SELECT id,action_no,action_type,full_name,final_status,created_at FROM housing_actions ORDER BY id DESC LIMIT 8").fetchall()
  urgent_tickets=c.execute("SELECT id,ticket_no,location_id,priority,status,created_at FROM maintenance_tickets WHERE priority IN ('urgent','critical') AND status!='closed' ORDER BY CASE priority WHEN 'critical' THEN 0 ELSE 1 END,id DESC LIMIT 8").fetchall()
  zone_rows=[]
  zone_names=sorted({str(x['zone'] or '-') for x in residential}, key=lambda z:(0,int(z)) if z.isdigit() else (1,z))
  for z in zone_names:
   zr=[x for x in residential if str(x['zone'] or '-')==z]
   zrooms=len(zr);zworkers=sum(x['occupied'] for x in zr);zcap=sum(x['capacity'] or 0 for x in zr);zfree=max(zcap-zworkers,0)
   zone_rows.append({'zone':z,'rooms':zrooms,'workers':zworkers,'capacity':zcap,'free':zfree,'pct':round(zworkers*100/zcap,1) if zcap else 0,'crowded':sum(1 for x in zr if x['occupied']>(x['capacity'] or 0))})
  supervisors=c.execute("SELECT id,display_name FROM users WHERE role='housing_supervisor' AND active=1 ORDER BY display_name").fetchall()
  sup_kpis=[]
  for sp in supervisors:
   total=c.execute('SELECT COUNT(*) FROM rooms WHERE supervisor_id=?',(sp['id'],)).fetchone()[0]
   insp=c.execute("SELECT COUNT(DISTINCT location_id) FROM inspections WHERE inspector_id=? AND inspection_type='room' AND week_key=?",(sp['id'],week)).fetchone()[0]
   sup_kpis.append({'name':sp['display_name'],'done':insp,'total':total,'pct':round(insp*100/total,1) if total else 0})
 total_rooms=max(rooms,1)
 occupancy_rate=round(workers*100/capacity,1) if capacity else 0
 free_beds=max(capacity-workers,0)
 ticket_total=open_t+closed_t
 ticket_close_rate=round(closed_t*100/ticket_total,1) if ticket_total else 0
 pct={'occupied':round(occupied*100/total_rooms,1),'vacant':round(vacant*100/total_rooms,1),'crowded':round(crowded*100/total_rooms,1),'closed':round(closed*100/max(len(room_stats),1),1),'maintenance':round(maintenance*100/max(len(room_stats),1),1)}
 title='لوحة مشرف السكن' if u['role']=='housing_supervisor' else 'لوحة مدير السكن والخدمات المساندة' if is_admin(u) else 'لوحة التحكم'
 return page('''<h2>{{dash_title}}</h2><p class="muted">واجهة اطلاع كاملة ومباشرة حسب صلاحيات المستخدم.</p>
 <div class="cards"><a class="card" href="{{url_for('workers')}}" style="color:inherit;text-decoration:none"><div>إجمالي العمال</div><div class="num">{{workers}}</div><small class="muted">نسبة إشغال الأسرة {{occupancy_rate}}%</small></a><div class="card"><div>إجمالي الطاقة الاستيعابية</div><div class="num">{{capacity}}</div><small class="muted">سرير</small></div><a class="card" href="{{url_for('available_beds')}}" style="color:inherit;text-decoration:none"><div>الأسرة الشاغرة</div><div class="num">{{free_beds}}</div><small class="muted">اضغط لعرض الغرف المتاحة</small></a><a class="card" href="{{url_for('tickets')}}" style="color:inherit;text-decoration:none"><div>بلاغات الصيانة المفتوحة</div><div class="num">{{open_t}}</div><small class="muted">نسبة الإغلاق {{ticket_close_rate}}%</small></a><a class="card" href="{{url_for('tickets')}}" style="color:inherit;text-decoration:none"><div>بلاغات الصيانة المغلقة</div><div class="num">{{closed_t}}</div></a><a class="card" href="{{url_for('rooms',usage='residential')}}" style="color:inherit;text-decoration:none"><div>الغرف السكنية</div><div class="num">{{rooms}}</div></a><a class="card" href="{{url_for('rooms',usage='warehouse')}}" style="color:inherit;text-decoration:none"><div>المستودعات</div><div class="num">{{usage_counts.warehouse}}</div></a><a class="card" href="{{url_for('rooms',usage='contractor')}}" style="color:inherit;text-decoration:none"><div>غرف المقاول</div><div class="num">{{usage_counts.contractor}}</div></a><a class="card" href="{{url_for('rooms',usage='security')}}" style="color:inherit;text-decoration:none"><div>غرف الحراسات الداخلية</div><div class="num">{{usage_counts.security}}</div></a><a class="card" href="{{url_for('occupancy_management',state='vacant')}}" style="color:inherit;text-decoration:none"><div>الغرف الفارغة</div><div class="num">{{vacant}}</div></a><a class="card" href="{{url_for('occupancy_management',state='maintenance')}}" style="color:inherit;text-decoration:none"><div>تحت الصيانة</div><div class="num">{{maintenance}}</div></a><a class="card" href="{{url_for('occupancy_management',state='closed')}}" style="color:inherit;text-decoration:none"><div>الغرف المغلقة</div><div class="num">{{closed}}</div></a><div class="card"><div>مجمعات دورات المياه</div><div class="num">{{bathroom_complexes_count}}</div></div><div class="card"><div>دورات المياه المكلف بها</div><div class="num">{{toilets_count}}</div></div><div class="card"><div>المغاسل المكلف بها</div><div class="num">{{basins_count}}</div></div><a class="card" href="{{url_for('absence_reports_list')}}" style="color:inherit;text-decoration:none"><div>بلاغات عدم التواجد</div><div class="num">{{absence_open}}</div></a><div class="card"><div>طلبات سكن معلقة</div><div class="num">{{pending_housing}}</div></div><div class="card"><div>إنجاز الجولات</div><div class="num">{{progress}}%</div></div></div>
 <div class="donuts"><div class="donut-card"><b>الإشغال</b><div class="donut" style="--p:{{pct.occupied}};--c:#25845f" data-label="{{pct.occupied}}%"></div></div><div class="donut-card"><b>الشواغر</b><div class="donut" style="--p:{{pct.vacant}};--c:#d49a18" data-label="{{pct.vacant}}%"></div></div><div class="donut-card"><b>التكدس</b><div class="donut" style="--p:{{pct.crowded}};--c:#b83232" data-label="{{pct.crowded}}%"></div></div><div class="donut-card"><b>المغلقة</b><div class="donut" style="--p:{{pct.closed}};--c:#222" data-label="{{pct.closed}}%"></div></div><div class="donut-card"><b>تحت الصيانة</b><div class="donut" style="--p:{{pct.maintenance}};--c:#2b6fb3" data-label="{{pct.maintenance}}%"></div></div></div>
 <div class="card" style="margin-top:14px"><h3>ملخص الإشغال حسب الزون</h3><div class="tbl-wrap"><table class="tbl"><tr><th>الزون</th><th>الغرف</th><th>العمال</th><th>الطاقة</th><th>الشاغر</th><th>الإشغال</th><th>التكدس</th></tr>{% for z in zone_rows %}<tr><td><b>{{z.zone}}</b></td><td>{{z.rooms}}</td><td>{{z.workers}}</td><td>{{z.capacity}}</td><td>{{z.free}}</td><td><div style="min-width:150px;background:#e7ecea;border-radius:10px;overflow:hidden"><div style="width:{{[z.pct,100]|min}}%;background:#25845f;color:white;padding:4px;text-align:center">{{z.pct}}%</div></div></td><td>{% if z.crowded %}<span class="badge red">{{z.crowded}}</span>{% else %}<span class="badge">0</span>{% endif %}</td></tr>{% endfor %}</table></div></div>
 <div class="grid" style="margin-top:14px"><div class="card"><h3>أحدث بلاغات الصيانة</h3><table class="tbl"><tr><th>البلاغ</th><th>الموقع</th><th>الحالة</th></tr>{% for x in latest_tickets %}<tr><td><a href="{{url_for('ticket_detail',tid=x.id)}}">{{x.ticket_no}}</a></td><td>{{x.location_id}}</td><td>{{status_ar.get(x.status,x.status)}}</td></tr>{% endfor %}</table></div><div class="card"><h3>أحدث طلبات السكن</h3><table class="tbl"><tr><th>الطلب</th><th>العامل</th><th>الحالة</th></tr>{% for x in latest_actions %}<tr><td><a href="{{url_for('worker_change_request_detail',qid=x.id)}}">{{x.action_no}}</a></td><td>{{x.full_name or '-'}}</td><td>{{x.final_status}}</td></tr>{% endfor %}</table></div></div>
 {% if urgent_tickets %}<div class="card"><h3>تنبيهات الصيانة العاجلة</h3><div class="tbl-wrap"><table class="tbl"><tr><th>البلاغ</th><th>الموقع</th><th>الأولوية</th><th>الحالة</th><th>التاريخ</th></tr>{% for x in urgent_tickets %}<tr><td><a href="{{url_for('ticket_detail',tid=x.id)}}">{{x.ticket_no}}</a></td><td>{{x.location_id}}</td><td><span class="badge red">{{'طارئ' if x.priority=='critical' else 'عاجل'}}</span></td><td>{{status_ar.get(x.status,x.status)}}</td><td>{{x.created_at}}</td></tr>{% endfor %}</table></div></div>{% endif %}
 {% if admin %}<div class="card"><h3>نسب إنجاز المشرفين لهذا الأسبوع</h3><div class="tbl-wrap"><table class="tbl"><tr><th>المشرف</th><th>الغرف المكلف بها</th><th>الغرف المنجزة</th><th>نسبة الإنجاز</th></tr>{% for x in sup_kpis %}<tr><td>{{x.name}}</td><td>{{x.total}}</td><td>{{x.done}}</td><td><b>{{x.pct}}%</b></td></tr>{% endfor %}</table></div></div>{% endif %}''','لوحة التحكم',u,dash_title=title,workers=workers,rooms=rooms,capacity=capacity,open_t=open_t,closed_t=closed_t,pending_housing=pending_housing,progress=progress,pct=pct,usage_counts=usage_counts,vacant=vacant,maintenance=maintenance,closed=closed,bathroom_complexes_count=bathroom_complexes_count,toilets_count=toilets_count,basins_count=basins_count,absence_open=absence_open,latest_tickets=latest_tickets,latest_actions=latest_actions,urgent_tickets=urgent_tickets,zone_rows=zone_rows,sup_kpis=sup_kpis,status_ar=STATUS_AR,admin=is_admin(u),occupancy_rate=occupancy_rate,free_beds=free_beds,ticket_close_rate=ticket_close_rate)

@app.get('/maintenance-dashboard')
@login_required
def maintenance_dashboard():
 u=current_user()
 if u['role'] not in ('super_admin','maintenance_manager','maintenance_supervisor','housing_manager','services_manager'):abort(403)
 where='';params=[]
 if u['role']=='maintenance_supervisor':where='WHERE assigned_to=? OR assigned_to IS NULL';params=[u['id']]
 with closing(conn()) as c:
  rows=c.execute(f'''SELECT t.*,u.display_name reporter FROM maintenance_tickets t LEFT JOIN users u ON u.id=t.reported_by {where} ORDER BY t.id DESC''',params).fetchall()
 cols=[('pending_maintenance','بانتظار القبول'),('accepted','مقبولة'),('in_progress','قيد التنفيذ'),('awaiting_reporter','بانتظار اعتماد المبلغ'),('closed','مغلقة')]
 groups={k:[x for x in rows if x['status']==k or (k=='pending_maintenance' and x['status']=='new')] for k,_ in cols}
 return page('''<h2>لوحة مدير الصيانة — التشغيلية</h2><p class="muted">البلاغ يمر بالقبول ثم التنفيذ ورفع صور الإغلاق ثم اعتماد مقدم البلاغ.</p><div class="kanban">{% for key,label in cols %}<div class="kanban-col"><h3>{{label}} ({{groups[key]|length}})</h3>{% for x in groups[key] %}<a class="ticket-card" style="display:block;color:inherit;text-decoration:none" href="{{url_for('ticket_detail',tid=x.id)}}"><b>{{x.ticket_no}}</b><div>{{x.location_type}} {{x.location_id}}</div><small>{{x.category}} · {{x.reporter}}</small></a>{% endfor %}</div>{% endfor %}</div>''','لوحة الصيانة',u,cols=cols,groups=groups)

@app.get('/search')
@login_required
def global_search():
 u=current_user();q=request.args.get('q','').strip();workers=[];rooms=[];tickets=[];absence=[]
 if q:
  with closing(conn()) as c:
   cl,args=assigned_clause(u,'r')
   workers=c.execute(f'''SELECT w.id,w.employee_no,w.full_name,w.room_no,w.zone FROM workers w JOIN rooms r ON r.room_no=w.room_no WHERE w.archived=0 AND {cl} AND (w.employee_no LIKE ? OR w.full_name LIKE ? OR w.iqama_no LIKE ? OR w.room_no LIKE ?) ORDER BY w.full_name LIMIT 50''',list(args)+[f'%{q}%']*4).fetchall()
   rooms=c.execute(f'''SELECT r.room_no,r.zone,r.usage_type,r.status FROM rooms r WHERE {cl} AND r.room_no LIKE ? ORDER BY CAST(r.room_no AS INTEGER) LIMIT 50''',list(args)+[f'%{q}%']).fetchall()
   if is_admin(u) or u['role'] in ('maintenance_manager','maintenance_supervisor'):
    tickets=c.execute("SELECT id,ticket_no,location_id,status FROM maintenance_tickets WHERE ticket_no LIKE ? OR location_id LIKE ? OR description LIKE ? ORDER BY id DESC LIMIT 50",[f'%{q}%']*3).fetchall()
   else:
    tickets=c.execute("SELECT id,ticket_no,location_id,status FROM maintenance_tickets WHERE reported_by=? AND (ticket_no LIKE ? OR location_id LIKE ? OR description LIKE ?) ORDER BY id DESC LIMIT 50",[u['id']]+[f'%{q}%']*3).fetchall()
   if is_admin(u):
    absence=c.execute('''SELECT a.id,a.room_no,a.status,w.employee_no,w.full_name FROM absence_reports a JOIN workers w ON w.id=a.worker_id WHERE w.employee_no LIKE ? OR w.full_name LIKE ? OR a.room_no LIKE ? ORDER BY a.id DESC LIMIT 50''',[f'%{q}%']*3).fetchall()
   elif u['role']=='housing_supervisor':
    absence=c.execute('''SELECT a.id,a.room_no,a.status,w.employee_no,w.full_name FROM absence_reports a JOIN workers w ON w.id=a.worker_id WHERE a.reported_by=? AND (w.employee_no LIKE ? OR w.full_name LIKE ? OR a.room_no LIKE ?) ORDER BY a.id DESC LIMIT 50''',[u['id']]+[f'%{q}%']*3).fetchall()
 return page('''<h2>البحث الشامل</h2><form class="card"><div class="grid"><div class="field"><label>العامل، الرقم الوظيفي، الإقامة، الغرفة أو البلاغ</label><input name="q" value="{{q}}" placeholder="اكتب كلمة البحث" autofocus></div></div><button class="btn">بحث</button></form>{% if q %}<div class="grid"><div class="card"><h3>العمال ({{workers|length}})</h3>{% for x in workers %}<p><a href="{{url_for('worker_detail',wid=x.id)}}"><b>{{x.employee_no}}</b> — {{x.full_name}} — غرفة {{x.room_no}}</a></p>{% else %}<p class="muted">لا توجد نتائج</p>{% endfor %}</div><div class="card"><h3>الغرف ({{rooms|length}})</h3>{% for x in rooms %}<p><a href="{{url_for('room_detail',room_no=x.room_no)}}">غرفة {{x.room_no}} — زون {{x.zone}} — {{usage_ar.get(x.usage_type,x.usage_type)}}</a></p>{% else %}<p class="muted">لا توجد نتائج</p>{% endfor %}</div><div class="card"><h3>بلاغات الصيانة ({{tickets|length}})</h3>{% for x in tickets %}<p><a href="{{url_for('ticket_detail',tid=x.id)}}">{{x.ticket_no}} — {{x.location_id}} — {{status_ar.get(x.status,x.status)}}</a></p>{% else %}<p class="muted">لا توجد نتائج</p>{% endfor %}</div><div class="card"><h3>بلاغات عدم التواجد ({{absence|length}})</h3>{% for x in absence %}<p><a href="{{url_for('absence_reports_list')}}">{{x.employee_no}} — {{x.full_name}} — غرفة {{x.room_no}} — {{x.status}}</a></p>{% else %}<p class="muted">لا توجد نتائج</p>{% endfor %}</div></div>{% endif %}''','البحث الشامل',u,q=q,workers=workers,rooms=rooms,tickets=tickets,absence=absence,usage_ar=ROOM_USAGE_AR,status_ar=STATUS_AR)

@app.get('/workers')
@login_required
def workers():
 u=current_user();q=request.args.get('q','').strip();zone=request.args.get('zone','').strip();status=request.args.get('status','').strip();cl,args=assigned_clause(u,'r')
 sql=f"""SELECT w.*,r.sector_name,u.display_name supervisor_name FROM workers w JOIN rooms r ON r.room_no=w.room_no LEFT JOIN users u ON u.id=r.supervisor_id WHERE w.archived=0 AND {cl}""";p=list(args)
 if q:
  like=f'%{q}%';sql+=' AND (w.employee_no LIKE ? OR w.iqama_no LIKE ? OR w.full_name LIKE ? OR w.room_no LIKE ? OR w.nationality LIKE ? OR w.profession LIKE ?)';p += [like]*6
 if zone:sql+=' AND CAST(w.zone AS TEXT)=?';p.append(zone)
 if status:sql+=" AND COALESCE(w.status,'active')=?";p.append(status)
 sql+=' ORDER BY CAST(w.zone AS INTEGER),CAST(w.room_no AS INTEGER),w.full_name LIMIT 1000'
 with closing(conn()) as c:
  rows=c.execute(sql,p).fetchall();zones=c.execute('SELECT DISTINCT zone FROM workers WHERE archived=0 AND zone IS NOT NULL ORDER BY CAST(zone AS INTEGER)').fetchall();total=c.execute('SELECT COUNT(*) FROM workers WHERE archived=0').fetchone()[0]
 return page('''<div class="card"><h2>بيانات العمالة</h2><form method="get" class="grid"><div class="field"><label>بحث ذكي</label><input name="q" value="{{q}}" placeholder="الرقم الوظيفي، الإقامة، الاسم، الغرفة، الجنسية، المهنة"></div><div class="field"><label>الزون</label><select name="zone"><option value="">الكل</option>{% for z in zones %}<option value="{{z.zone}}" {% if zone==z.zone|string %}selected{% endif %}>{{z.zone}}</option>{% endfor %}</select></div><div class="field"><label>الحالة</label><select name="status"><option value="">الكل</option><option value="active" {% if status=='active' %}selected{% endif %}>نشط</option><option value="outside_temp" {% if status=='outside_temp' %}selected{% endif %}>سكن خارجي مؤقت</option><option value="outside_perm" {% if status=='outside_perm' %}selected{% endif %}>سكن خارجي دائم</option><option value="temporary_exit" {% if status=='temporary_exit' %}selected{% endif %}>خروج مؤقت</option></select></div><div class="field"><label>&nbsp;</label><button class="btn">بحث</button></div></form><p class="muted">المعروض {{rows|length}} من إجمالي {{total}} عامل</p>{% if export_allowed %}<a class="btn" href="{{url_for('export_workers_excel',q=q,zone=zone,status=status)}}">تصدير Excel</a> <a class="btn btn2" href="{{url_for('export_workers_pdf',q=q,zone=zone,status=status)}}">تصدير PDF</a> <button class="btn btn2" onclick="window.print()">طباعة</button>{% endif %}</div><div class="tbl-wrap"><table class="tbl"><tr><th>الرقم الوظيفي</th><th>رقم الإقامة</th><th>الاسم</th><th>الجنسية</th><th>المهنة</th><th>الهاتف</th><th>الزون</th><th>الغرفة</th><th>القطاع</th><th>المشرف</th><th>الحالة</th><th></th></tr>{% for w in rows %}<tr><td>{{w.employee_no}}</td><td>{{w.iqama_no or '-'}}</td><td>{{w.full_name}}</td><td>{{w.nationality}}</td><td>{{w.profession}}</td><td>{{w.phone or '-'}}</td><td>{{w.zone}}</td><td>{{w.room_no}}</td><td>{{w.sector_name or '-'}}</td><td>{{w.supervisor_name or '-'}}</td><td>{{w.status}}</td><td><a href="{{url_for('worker_detail',wid=w.id)}}">فتح</a></td></tr>{% else %}<tr><td colspan="12">لا توجد نتائج</td></tr>{% endfor %}</table></div>''','العمال',u,rows=rows,q=q,zone=zone,status=status,zones=zones,total=total,export_allowed=can_export(u))

@app.get('/workers/<int:wid>')
@login_required
def worker_detail(wid):
 u=current_user();cl,args=assigned_clause(u,'r')
 with closing(conn()) as c:
  w=c.execute(f'''SELECT w.* FROM workers w LEFT JOIN rooms r ON r.room_no=w.room_no WHERE w.id=? AND COALESCE(w.archived,0)=0 AND {cl}''',[wid]+list(args)).fetchone()
  if not w:abort(404)
  events=c.execute('''SELECT e.*,u.display_name FROM worker_events e LEFT JOIN users u ON u.id=e.user_id WHERE e.worker_id=? ORDER BY e.id DESC LIMIT 50''',(wid,)).fetchall()
  absences=c.execute('''SELECT a.*,u.display_name reporter FROM absence_reports a LEFT JOIN users u ON u.id=a.reported_by WHERE a.worker_id=? ORDER BY a.id DESC''',(wid,)).fetchall()
 return page('''<div class="card"><h2>{{w.full_name}}</h2><div class="grid"><p><b>الرقم الوظيفي:</b> {{w.employee_no}}</p><p><b>رقم الإقامة:</b> {{w.iqama_no or '-'}}</p><p><b>الجنسية:</b> {{w.nationality or '-'}}</p><p><b>المهنة:</b> {{w.profession or '-'}}</p><p><b>الجوال:</b> {{w.phone or '-'}}</p><p><b>الزون:</b> {{w.zone or '-'}}</p><p><b>الغرفة:</b> <a href="{{url_for('room_detail',room_no=w.room_no)}}">{{w.room_no}}</a></p></div>{% if u.role=='housing_supervisor' %}<a class="btn" href="{{url_for('new_absence_report',wid=w.id)}}">رفع بلاغ عدم تواجد</a>{% endif %} <a class="btn btn2" href="{{url_for('housing_kit_worker',wid=w.id)}}">العهدة السكنية</a></div><div class="card"><h3>بلاغات عدم التواجد</h3><div class="tbl-wrap"><table class="tbl"><tr><th>التاريخ</th><th>المدة</th><th>الحالة</th><th>المشرف</th><th>الملاحظات</th></tr>{% for x in absences %}<tr><td>{{x.created_at}}</td><td>{{x.absence_duration}}</td><td>{{x.status}}</td><td>{{x.reporter or '-'}}</td><td>{{x.notes}}</td></tr>{% else %}<tr><td colspan="5" class="muted">لا توجد بلاغات مسجلة لهذا العامل.</td></tr>{% endfor %}</table></div></div><div class="card"><h3>سجل العامل</h3><div class="timeline">{% for x in events %}<div class="timeline-item"><b>{{x.event_type}}</b><p>{{x.description or '-'}}</p><small>{{x.created_at}} - {{x.display_name or '-'}}</small></div>{% else %}<p class="muted">لا يوجد سجل عمليات حتى الآن.</p>{% endfor %}</div></div>''','تفاصيل العامل',u,w=w,events=events,absences=absences,u=u)

@app.get('/rooms')
@login_required
def rooms():
 u=current_user();q=request.args.get('q','').strip();usage=request.args.get('usage','').strip();cl,args=assigned_clause(u,'r');sql=f'''SELECT r.*,COUNT(w.id) occupied FROM rooms r LEFT JOIN workers w ON w.room_no=r.room_no AND w.archived=0 WHERE {cl}''';p=list(args)
 if q:sql+=' AND r.room_no LIKE ?';p.append(f'%{q}%')
 if usage:sql+=" AND COALESCE(r.usage_type,'residential')=?";p.append(usage)
 sql+=' GROUP BY r.id ORDER BY CAST(r.room_no AS INTEGER),r.room_no LIMIT 1000'
 with closing(conn()) as c:rows=c.execute(sql,p).fetchall()
 return page('''<h2>إدارة الغرف</h2><form class="card"><div class="grid"><div class="field"><label>رقم الغرفة</label><input name="q" value="{{q}}" placeholder="بحث"></div><div class="field"><label>الاستخدام</label><select name="usage"><option value="">الكل</option>{% for k,v in usage_ar.items() %}<option value="{{k}}" {% if usage==k %}selected{% endif %}>{{v}}</option>{% endfor %}</select></div></div><button class="btn">تصفية</button></form><div class="tbl-wrap"><table class="tbl"><tr><th>الزون</th><th>الغرفة</th><th>الاستخدام</th><th>المساحة</th><th>السعة</th><th>المشغول</th><th>الشاغر</th><th>الحالة</th><th></th></tr>{% for x in rows %}{% set free=x.capacity-x.occupied %}<tr><td>{{x.zone}}</td><td><b>{{x.room_no}}</b></td><td>{{usage_ar.get(x.usage_type or 'residential',x.usage_type)}}</td><td>{{'%.2f'|format(x.area_m2) if x.area_m2 else '-'}}</td><td>{{x.capacity}}</td><td>{{x.occupied}}</td><td>{{free}}</td><td>{% if x.occupied>x.capacity %}<span class="badge red">متكدسة</span>{% elif x.occupied==0 and (x.usage_type or 'residential')=='residential' %}<span class="badge amber">فارغة</span>{% else %}<span class="badge">طبيعية</span>{% endif %}</td><td><a href="{{url_for('room_detail',room_no=x.room_no)}}">فتح</a></td></tr>{% endfor %}</table></div>''','الغرف',u,rows=rows,q=q,usage=usage,usage_ar=ROOM_USAGE_AR)

@app.route('/rooms/<room_no>',methods=['GET','POST'])
@login_required
def room_detail(room_no):
 u=current_user();cl,args=assigned_clause(u,'r')
 with closing(conn()) as c:
  r=c.execute(f'SELECT r.* FROM rooms r WHERE r.room_no=? AND {cl}',[room_no]+list(args)).fetchone()
  if not r:abort(404)
  if request.method=='POST':
   if not is_admin(u):abort(403)
   old=r['usage_type'] or 'residential';new=request.form.get('usage_type','residential')
   length=float(request.form['length_m']) if request.form.get('length_m') else None;width=float(request.form['width_m']) if request.form.get('width_m') else None
   area=round(length*width,2) if length and width else None
   capacity=int(request.form.get('capacity') or (max(1,int(area//4)) if area else r['capacity']))
   c.execute('UPDATE rooms SET usage_type=?,length_m=?,width_m=?,area_m2=?,capacity=?,status=?,notes=?,updated_at=? WHERE id=?',(new,length,width,area,capacity,request.form.get('status','active'),request.form.get('notes',''),now(),r['id']))
   if old!=new:c.execute('INSERT INTO room_usage_history(room_id,old_usage,new_usage,reason,changed_by,changed_at) VALUES(?,?,?,?,?,?)',(r['id'],old,new,request.form.get('reason',''),u['id'],now()))
   audit(c,u,'update','room',r['id'],{'room_no':room_no,'usage':new,'capacity':capacity});c.commit()
   return redirect(url_for('room_detail',room_no=room_no))
  workers=c.execute('SELECT * FROM workers WHERE room_no=? AND archived=0 ORDER BY full_name',(room_no,)).fetchall()
  tickets=c.execute("SELECT * FROM maintenance_tickets WHERE location_type='room' AND location_id=? ORDER BY id DESC",(room_no,)).fetchall()
  inspections_history=c.execute("SELECT * FROM inspections WHERE inspection_type='room' AND location_id=? ORDER BY id DESC",(room_no,)).fetchall()
  history=c.execute('''SELECT h.*,u.display_name FROM room_usage_history h LEFT JOIN users u ON u.id=h.changed_by WHERE h.room_id=? ORDER BY h.id DESC''',(r['id'],)).fetchall()
 return page('''<div class="card"><h2>الغرفة {{r.room_no}}</h2><div class="cards"><div><b>الزون</b><div class="num">{{r.zone}}</div></div><div><b>السعة</b><div class="num">{{r.capacity}}</div></div><div><b>المقيمون</b><div class="num">{{workers|length}}</div></div><div><b>الشاغر</b><div class="num">{{r.capacity-(workers|length)}}</div></div></div></div>
 {% if admin %}<div class="card"><h3>بيانات واستخدام الغرفة</h3><form method="post"><div class="grid"><div class="field"><label>الاستخدام</label><select name="usage_type">{% for k,v in usage_ar.items() %}<option value="{{k}}" {% if (r.usage_type or 'residential')==k %}selected{% endif %}>{{v}}</option>{% endfor %}</select></div><div class="field"><label>الطول بالمتر</label><input type="number" step="0.01" name="length_m" value="{{r.length_m or ''}}"></div><div class="field"><label>العرض بالمتر</label><input type="number" step="0.01" name="width_m" value="{{r.width_m or ''}}"></div><div class="field"><label>السعة المعتمدة</label><input type="number" min="0" name="capacity" value="{{r.capacity}}"></div><div class="field"><label>الحالة</label><select name="status"><option value="active">نشطة</option><option value="closed">مغلقة</option><option value="out_of_service">خارج الخدمة</option></select></div><div class="field"><label>سبب تغيير الاستخدام</label><input name="reason"></div></div><div class="field"><label>ملاحظات</label><textarea name="notes">{{r.notes or ''}}</textarea></div><button class="btn">حفظ التعديلات</button></form></div>{% endif %}
 <div class="card"><h3>المقيمون</h3><div class="tbl-wrap"><table class="tbl"><tr><th>الرقم</th><th>الاسم</th><th>الجنسية</th><th>المهنة</th></tr>{% for w in workers %}<tr><td>{{w.employee_no}}</td><td>{{w.full_name}}</td><td>{{w.nationality}}</td><td>{{w.profession}}</td></tr>{% endfor %}</table></div></div>
 <div class="card"><a class="btn" href="{{url_for('new_inspection',room_no=r.room_no)}}">تنفيذ الجولة الأسبوعية</a> <a class="btn btn2" href="{{url_for('new_ticket')}}?location_type=room&location_id={{r.room_no}}">رفع بلاغ صيانة</a></div><div class="grid"><div class="card"><h3>بلاغات الصيانة</h3>{% for x in tickets %}<p><a href="{{url_for('ticket_detail',tid=x.id)}}">{{x.ticket_no}}</a> · {{x.category}} · {{x.status}}</p>{% else %}<p class="muted">لا توجد بلاغات</p>{% endfor %}</div><div class="card"><h3>سجل الجولات</h3>{% for x in inspections_history %}<p>{{x.created_at}} · {{x.cleanliness or '-'}} · {{x.notes or ''}}</p>{% else %}<p class="muted">لا توجد جولات</p>{% endfor %}</div></div>
 {% if history %}<div class="card"><h3>سجل تغيير الاستخدام</h3><div class="tbl-wrap"><table class="tbl"><tr><th>التاريخ</th><th>السابق</th><th>الجديد</th><th>السبب</th><th>بواسطة</th></tr>{% for h in history %}<tr><td>{{h.changed_at}}</td><td>{{usage_ar.get(h.old_usage,h.old_usage)}}</td><td>{{usage_ar.get(h.new_usage,h.new_usage)}}</td><td>{{h.reason}}</td><td>{{h.display_name}}</td></tr>{% endfor %}</table></div></div>{% endif %}''','الغرفة '+room_no,u,r=r,workers=workers,history=history,tickets=tickets,inspections_history=inspections_history,usage_ar=ROOM_USAGE_AR,admin=is_admin(u))
@app.get('/inspections')
@login_required
def inspections():
 u=current_user();
 if u['role']!='housing_supervisor' and not is_admin(u):abort(403)
 week=date.today().strftime('%Y-W%W');cl,args=assigned_clause(u,'r')
 with closing(conn()) as c:
  room_rows=c.execute(f'''SELECT r.room_no,r.zone,r.capacity,COUNT(DISTINCT w.id) occupied,
   GROUP_CONCAT(DISTINCT CASE WHEN w.id IS NOT NULL THEN w.employee_no||' - '||w.full_name END) worker_names,
   MAX(CASE WHEN i.week_key=? THEN i.id END) week_inspection_id,
   MAX(CASE WHEN i.week_key=? THEN i.created_at END) week_inspection_at,
   MAX(i.created_at) last_inspection_at
   FROM rooms r
   LEFT JOIN workers w ON w.room_no=r.room_no AND w.archived=0
   LEFT JOIN inspections i ON i.inspection_type='room' AND i.location_id=r.room_no
   WHERE {cl} AND COALESCE(r.usage_type,'residential')='residential'
   GROUP BY r.id ORDER BY CAST(r.room_no AS INTEGER),r.room_no''',[week,week]+list(args)).fetchall()
  if u['role']=='housing_supervisor':
   history=c.execute("SELECT i.*,u.display_name,t.ticket_no FROM inspections i JOIN users u ON u.id=i.inspector_id LEFT JOIN maintenance_tickets t ON t.id=i.maintenance_ticket_id WHERE i.inspection_type='room' AND i.inspector_id=? ORDER BY i.id DESC LIMIT 300",(u['id'],)).fetchall()
  else:
   history=c.execute("SELECT i.*,u.display_name,t.ticket_no FROM inspections i JOIN users u ON u.id=i.inspector_id LEFT JOIN maintenance_tickets t ON t.id=i.maintenance_ticket_id WHERE i.inspection_type='room' ORDER BY i.id DESC LIMIT 300").fetchall()
 return page('''<h2>الجولة الأسبوعية للغرف</h2><p class="muted">تظهر لك الغرف الواقعة ضمن نطاق صلاحيتك مع أسماء العمال المسجلين داخل كل غرفة.</p>
 <div class="cards"><div class="card"><div>الغرف المكلف بها</div><div class="num">{{room_rows|length}}</div></div><div class="card"><div>تمت جولتها هذا الأسبوع</div><div class="num">{{room_rows|selectattr('week_inspection_id')|list|length}}</div></div><div class="card"><div>متبقي</div><div class="num">{{room_rows|rejectattr('week_inspection_id')|list|length}}</div></div></div>
 <div class="tbl-wrap"><table class="tbl"><tr><th>الزون</th><th>الغرفة</th><th>السعة</th><th>المسجلون</th><th>أسماء العمال</th><th>حالة الجولة</th><th></th></tr>{% for r in room_rows %}<tr><td>{{r.zone}}</td><td><b>{{r.room_no}}</b></td><td>{{r.capacity}}</td><td>{{r.occupied}}</td><td style="white-space:normal;min-width:300px">{{(r.worker_names or 'لا يوجد عمال مسجلون')|replace(',', ' | ')}}</td><td>{% if r.week_inspection_id %}<span class="badge">تمت {{r.week_inspection_at}}</span>{% else %}<span class="badge amber">لم تتم</span>{% endif %}</td><td><a class="btn" href="{{url_for('new_inspection',room_no=r.room_no)}}">{{'إعادة الجولة' if r.week_inspection_id else 'بدء الجولة'}}</a></td></tr>{% endfor %}</table></div>
 {% if history %}<div class="card"><h3>سجل الجولات السابقة</h3><div class="tbl-wrap"><table class="tbl"><tr><th>التاريخ</th><th>الغرفة</th><th>المسجل</th><th>الفعلي</th><th>النظافة</th><th>المشرف</th><th>بلاغ الصيانة</th><th>ملاحظات</th></tr>{% for x in history %}<tr><td>{{x.created_at}}</td><td>{{x.location_id}}</td><td>{{x.registered_count}}</td><td>{{x.actual_count}}</td><td>{{x.cleanliness}}</td><td>{{x.display_name}}</td><td>{{x.ticket_no or '-'}}</td><td>{{x.notes}}</td></tr>{% endfor %}</table></div></div>{% endif %}''','الجولة الأسبوعية',u,room_rows=room_rows,history=history)

@app.route('/inspections/new',methods=['GET','POST'])
@login_required
def new_inspection():
 u=current_user();
 if u['role']!='housing_supervisor':abort(403)
 room=(request.args.get('room_no') or request.form.get('room_no') or '').strip();err='';workers=[];room_row=None
 with closing(conn()) as c:
  cl,args=assigned_clause(u,'r')
  if room:
   room_row=c.execute(f'SELECT r.* FROM rooms r WHERE r.room_no=? AND {cl}',[room]+list(args)).fetchone()
   if room_row:workers=c.execute('SELECT employee_no,full_name,nationality,profession FROM workers WHERE room_no=? AND archived=0 ORDER BY full_name',(room,)).fetchall()
  if request.method=='POST':
   if not room_row:err='رقم الغرفة غير موجود أو ليس ضمن نطاق صلاحيتك'
   else:
    clean=request.form.get('cleanliness');actual=request.form.get('actual_count');notes=request.form.get('notes','');reg=len(workers);wk=date.today().strftime('%Y-W%W')
    cur=c.execute('INSERT INTO inspections(inspection_type,location_id,zone_name,inspector_id,registered_count,actual_count,cleanliness,notes,status,created_at,week_key) VALUES(?,?,?,?,?,?,?,?,?,?,?)',('room',room,str(room_row['zone']),u['id'],reg,int(actual or 0),clean,notes,'completed',now(),wk))
    inspection_id=cur.lastrowid;ticket_id=None
    if request.form.get('maintenance_required')=='1':
     try: photo=save_upload(request.files.get('maintenance_photo'),'inspection')
     except ValueError as e: photo=None; err=str(e)
     category=request.form.get('maintenance_category','أخرى');description=request.form.get('maintenance_description','').strip();priority=request.form.get('maintenance_priority','normal')
     if not description:
      c.rollback();err='اكتب وصف عطل الصيانة قبل حفظ الجولة'
     elif not photo:
      c.rollback();err='يجب إرفاق صورة لعطل الصيانة'
     else:
      ticket_no='MNT-'+datetime.utcnow().strftime('%Y%m%d%H%M%S%f')[-16:]
      t=c.execute('INSERT INTO maintenance_tickets(ticket_no,location_type,location_id,zone_name,category,description,priority,status,reported_by,before_photo,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(ticket_no,'room',room,str(room_row['zone']),category,description,priority,'new',u['id'],photo,now()))
      ticket_id=t.lastrowid;c.execute('UPDATE inspections SET maintenance_ticket_id=? WHERE id=?',(ticket_id,inspection_id));c.execute('INSERT INTO ticket_photos(ticket_id,photo_path,photo_kind,uploaded_by,created_at) VALUES(?,?,?,?,?)',(ticket_id,photo,'report',u['id'],now()));c.execute('INSERT INTO ticket_updates(ticket_id,user_id,action,notes,photo_path,created_at) VALUES(?,?,?,?,?,?)',(ticket_id,u['id'],'created_from_inspection',description,photo,now()));audit(c,u,'create','maintenance_ticket',ticket_id,{'source':'weekly_inspection','inspection_id':inspection_id,'room_no':room})
    if not err:
     audit(c,u,'create','inspection',inspection_id,{'room_no':room,'maintenance_ticket_id':ticket_id});c.commit();return redirect(url_for('inspections'))
 return page('''<div class="card"><h2>جولة الغرفة {{room or ''}}</h2>{% if err %}<p class="err">{{err}}</p>{% endif %}{% if room_row %}<div class="cards"><div><b>الزون</b><div class="num">{{room_row.zone}}</div></div><div><b>السعة</b><div class="num">{{room_row.capacity}}</div></div><div><b>العمال المسجلون</b><div class="num">{{workers|length}}</div></div></div><h3>أسماء العمال</h3><div class="tbl-wrap"><table class="tbl"><tr><th>الرقم</th><th>الاسم</th><th>الجنسية</th><th>المهنة</th></tr>{% for w in workers %}<tr><td>{{w.employee_no}}</td><td>{{w.full_name}}</td><td>{{w.nationality}}</td><td>{{w.profession}}</td></tr>{% endfor %}{% if not workers %}<tr><td colspan="4">لا يوجد عمال مسجلون في الغرفة</td></tr>{% endif %}</table></div>{% endif %}</div>
 <div class="card"><form method="post" enctype="multipart/form-data"><input type="hidden" name="room_no" value="{{room}}"><div class="grid"><div class="field"><label>العدد الفعلي</label><input type="number" min="0" name="actual_count" value="{{workers|length}}" required></div><div class="field"><label>النظافة</label><select name="cleanliness"><option>ممتاز</option><option>جيد</option><option>يحتاج متابعة</option></select></div></div><div class="field"><label>ملاحظات الجولة</label><textarea name="notes"></textarea></div>
 <div class="card" style="border:1px solid #d8dddd"><h3>الصيانة داخل الجولة</h3><label><input type="checkbox" name="maintenance_required" value="1" onchange="document.getElementById('maintenance-fields').style.display=this.checked?'grid':'none'"> يوجد عطل صيانة ويحتاج إنشاء بلاغ</label><div id="maintenance-fields" class="grid" style="display:none;margin-top:12px"><div class="field"><label>التصنيف</label><select name="maintenance_category"><option>كهرباء</option><option>سباكة</option><option>تكييف</option><option>نجارة</option><option>أثاث</option><option>أبواب ونوافذ</option><option>أخرى</option></select></div><div class="field"><label>الأولوية</label><select name="maintenance_priority"><option value="normal">عادي</option><option value="urgent">عاجل</option><option value="critical">طارئ</option></select></div><div class="field"><label>صورة العطل</label><input type="file" name="maintenance_photo" accept="image/*" capture="environment"></div><div class="field" style="grid-column:1/-1"><label>وصف العطل</label><textarea name="maintenance_description"></textarea></div></div></div><button class="btn">حفظ الجولة</button></form></div>''','جولة غرفة',u,room=room,err=err,room_row=room_row,workers=workers)

@app.get('/bathroom-inspections')
@login_required
def bathroom_inspections():
 return redirect(url_for('bathroom_reports'))

@app.get('/bathroom-reports')
@login_required
def bathroom_reports():
 # Phase 5: bathroom issues are handled only through maintenance.
 return redirect(url_for('new_ticket', location_type='bathroom', category='سباكة'))

@app.route('/bathroom-reports/new',methods=['GET','POST'])
@login_required
def new_bathroom_report():
 return redirect(url_for('new_ticket', location_type='bathroom', category='سباكة'))

@app.route('/requests/new',methods=['GET','POST'])
@login_required
def new_request():
 u=current_user()
 if u['role'] not in ('housing_monitor','housing_manager','services_manager'):abort(403)
 err='';worker=None
 if request.method=='POST':
  eno=request.form['employee_no'].strip();typ=request.form['request_type'];payload={k:request.form.get(k,'') for k in ('new_room','start_date','end_date','outside_location','reason')}
  with closing(conn()) as c:
   worker=c.execute('SELECT * FROM workers WHERE employee_no=? AND archived=0',(eno,)).fetchone()
   if not worker:err='العامل غير موجود أو مؤرشف'
   else:
    sup=supervisor_for_room(worker['room_no']);reqno='REQ-'+datetime.utcnow().strftime('%Y%m%d%H%M%S%f')[-16:]
    cur=c.execute('INSERT INTO requests(request_no,request_type,worker_id,payload_json,requested_by,approver_id,supervisor_id,status,created_at) VALUES(?,?,?,?,?,?,?,?,?)',(reqno,typ,worker['id'],json.dumps(payload,ensure_ascii=False),u['id'],sup['id'] if sup else None,sup['id'] if sup else None,'pending',now()));audit(c,u,'create','request',cur.lastrowid,{'type':typ,'worker':eno});c.commit();return redirect(url_for('request_detail',rid=cur.lastrowid))
 return page('''<div class="card"><h2>إنشاء طلب عامل</h2>{% if err %}<p class="err">{{err}}</p>{% endif %}<form method="post"><div class="grid"><div class="field"><label>الرقم الوظيفي للعامل</label><input name="employee_no" required></div><div class="field"><label>نوع الطلب</label><select name="request_type"><option value="transfer">نقل عامل</option><option value="final_exit">خروج نهائي</option><option value="outside_temp">سكن خارجي مؤقت</option><option value="outside_perm">سكن خارجي دائم</option></select></div><div class="field"><label>الغرفة الجديدة (للنقل)</label><input name="new_room"></div><div class="field"><label>تاريخ البداية</label><input type="date" name="start_date"></div><div class="field"><label>تاريخ النهاية</label><input type="date" name="end_date"></div><div class="field"><label>موقع السكن الخارجي</label><input name="outside_location"></div></div><div class="field"><label>السبب</label><textarea name="reason"></textarea></div><button class="btn">إرسال الطلب</button></form></div>''','إنشاء طلب',u,err=err)
@app.route('/requests/<int:rid>',methods=['GET','POST'])
@login_required
def request_detail(rid):
 u=current_user()
 with closing(conn()) as c:
  q=c.execute('''SELECT q.*,w.employee_no,w.full_name,w.room_no,u.display_name approver FROM requests q LEFT JOIN workers w ON w.id=q.worker_id LEFT JOIN users u ON u.id=q.supervisor_id WHERE q.id=?''',(rid,)).fetchone()
  if not q:abort(404)
  if u['role']=='housing_monitor' and q['requested_by']!=u['id']:abort(403)
  if u['role']=='housing_supervisor' and q['supervisor_id']!=u['id']:abort(403)
  if request.method=='POST':
   if not (is_admin(u) or (u['role']=='housing_supervisor' and q['supervisor_id']==u['id'])):abort(403)
   decision=request.form['decision'];reason=request.form.get('decision_reason','');payload=json.loads(q['payload_json'] or '{}')
   if q['status']!='pending':return redirect(url_for('request_detail',rid=rid))
   if decision=='approved':
    if q['request_type']=='transfer':
     new_room=payload.get('new_room','').strip();room=c.execute('SELECT * FROM rooms WHERE room_no=?',(new_room,)).fetchone()
     if not room:return page('<div class="card"><p class="err">الغرفة الجديدة غير موجودة.</p></div>','خطأ',u),400
     occupied=c.execute('SELECT COUNT(*) FROM workers WHERE room_no=? AND archived=0',(new_room,)).fetchone()[0]
     if occupied>=room['capacity']:return page('<div class="card"><p class="err">الغرفة الجديدة ممتلئة.</p></div>','خطأ',u),400
     c.execute("UPDATE workers SET room_no=?,zone=?,status='active',updated_at=? WHERE id=?",(new_room,room['zone'],now(),q['worker_id']))
    elif q['request_type']=='final_exit':c.execute("UPDATE workers SET archived=1,status='final_exit',updated_at=? WHERE id=?",(now(),q['worker_id']))
    elif q['request_type']=='outside_temp':c.execute("UPDATE workers SET status='outside_temp',updated_at=? WHERE id=?",(now(),q['worker_id']))
    elif q['request_type']=='outside_perm':c.execute("UPDATE workers SET status='outside_perm',updated_at=? WHERE id=?",(now(),q['worker_id']))
   c.execute('UPDATE requests SET status=?,approver_id=?,decision_reason=?,decided_at=?,supervisor_decision_at=? WHERE id=?',(decision,u['id'],reason,now(),now(),rid));audit(c,u,decision,'request',rid,{'reason':reason});c.commit();return redirect(url_for('request_detail',rid=rid))
  payload=json.loads(q['payload_json'] or '{}')
 return page('''<div class="card"><h2>طلب {{q.request_no}}</h2><p><b>النوع:</b> {{req_ar.get(q.request_type)}}</p><p><b>العامل:</b> {{q.employee_no}} - {{q.full_name}}</p><p><b>الغرفة الحالية:</b> {{q.room_no}}</p><p><b>المشرف المسؤول:</b> {{q.approver or '-'}}</p><p><b>الحالة:</b> {{status_ar.get(q.status,q.status)}}</p><p><b>البيانات:</b> {{payload}}</p>{% if can_decide and q.status=='pending' %}<form method="post"><textarea name="decision_reason" placeholder="ملاحظة القرار"></textarea><br><button class="btn" name="decision" value="approved">اعتماد وتنفيذ</button> <button class="btn danger" name="decision" value="rejected">رفض</button></form>{% endif %}</div>''','تفاصيل الطلب',u,q=q,payload=payload,req_ar=REQ_AR,status_ar=STATUS_AR,can_decide=is_admin(u) or (u['role']=='housing_supervisor' and q['supervisor_id']==u['id']))

@app.get('/worker-change-requests')
@login_required
def worker_change_requests():
 u=current_user()
 if not (can_create_housing(u) or is_admin(u)):abort(403)
 status=request.args.get('status','').strip();typ=request.args.get('type','').strip();qtxt=request.args.get('q','').strip()
 with closing(conn()) as c:
  where=[];params=[]
  if is_admin(u): pass
  elif u['role']=='housing_supervisor': where.append('(q.requested_by=? OR q.source_supervisor_id=? OR q.target_supervisor_id=?)');params += [u['id'],u['id'],u['id']]
  else: where.append('q.requested_by=?');params.append(u['id'])
  if status: where.append('q.final_status=?');params.append(status)
  if typ: where.append('q.action_type=?');params.append(typ)
  if qtxt:
   where.append('(q.action_no LIKE ? OR q.employee_no LIKE ? OR q.full_name LIKE ? OR q.source_room LIKE ? OR q.target_room LIKE ?)');like=f'%{qtxt}%';params += [like]*5
  sql='''SELECT q.*,ru.display_name requester FROM housing_actions q LEFT JOIN users ru ON ru.id=q.requested_by'''
  if where: sql+=' WHERE '+' AND '.join(where)
  sql+=' ORDER BY q.id DESC LIMIT 400'
  rows=c.execute(sql,params).fetchall()
  base_where=[];base_params=[]
  if not is_admin(u):
   if u['role']=='housing_supervisor':base_where.append('(requested_by=? OR source_supervisor_id=? OR target_supervisor_id=?)');base_params += [u['id'],u['id'],u['id']]
   else:base_where.append('requested_by=?');base_params.append(u['id'])
  bw=(' WHERE '+' AND '.join(base_where)) if base_where else ''
  counts={r['final_status']:r['n'] for r in c.execute('SELECT final_status,COUNT(*) n FROM housing_actions'+bw+' GROUP BY final_status',base_params).fetchall()}
 return page('''<h2>إدارة الطلبات</h2><p class="muted">وحدة تشغيلية للتسكين والنقل والخروج مع الاعتماد والمتابعة.</p><div class="cards"><a class="card" href="?" style="color:inherit;text-decoration:none"><div>الكل</div><div class="num">{{rows|length}}</div></a><a class="card" href="?status=pending_supervisors" style="color:inherit;text-decoration:none"><div>بانتظار المشرفين</div><div class="num">{{counts.get('pending_supervisors',0)}}</div></a><a class="card" href="?status=pending_management" style="color:inherit;text-decoration:none"><div>بانتظار الإدارة</div><div class="num">{{counts.get('pending_management',0)}}</div></a><a class="card" href="?status=approved" style="color:inherit;text-decoration:none"><div>معتمدة</div><div class="num">{{counts.get('approved',0)}}</div></a><a class="card" href="?status=rejected" style="color:inherit;text-decoration:none"><div>مرفوضة</div><div class="num">{{counts.get('rejected',0)}}</div></a></div>{% if creator %}<a class="btn" href="{{url_for('new_worker_change_request')}}">إنشاء طلب جديد</a>{% endif %}<form class="card" method="get"><div class="grid"><div class="field"><label>بحث</label><input name="q" value="{{qtxt}}" placeholder="رقم الطلب أو العامل أو الغرفة"></div><div class="field"><label>نوع الطلب</label><select name="type"><option value="">الكل</option>{% for k,v in types.items() %}<option value="{{k}}" {% if typ==k %}selected{% endif %}>{{v}}</option>{% endfor %}</select></div><div class="field"><label>الحالة</label><select name="status"><option value="">الكل</option>{% for k,v in statuses.items() %}<option value="{{k}}" {% if status==k %}selected{% endif %}>{{v}}</option>{% endfor %}</select></div></div><button class="btn">تصفية</button> <a class="btn btn2" href="{{url_for('worker_change_requests')}}">مسح</a></form><div class="tbl-wrap"><table class="tbl"><tr><th>الطلب</th><th>النوع</th><th>العامل</th><th>من</th><th>إلى</th><th>الحالة</th><th>مقدم الطلب</th><th>التاريخ</th><th>الإجراء</th></tr>{% for x in rows %}<tr><td>{{x.action_no}}</td><td>{{types.get(x.action_type,x.action_type)}}</td><td>{{x.employee_no or '-'}} - {{x.full_name or '-'}}</td><td>{{x.source_room or '-'}}</td><td>{{x.target_room or '-'}}</td><td><span class="badge">{{statuses.get(x.final_status,x.final_status)}}</span></td><td>{{x.requester or '-'}}</td><td>{{x.created_at}}</td><td><a class="btn" href="{{url_for('worker_change_request_detail',qid=x.id)}}">فتح الطلب</a></td></tr>{% else %}<tr><td colspan="9" class="muted">لا توجد طلبات مطابقة.</td></tr>{% endfor %}</table></div>''','إدارة الطلبات',u,rows=rows,creator=can_create_housing(u),qtxt=qtxt,typ=typ,status=status,counts=counts,types={'add':'تسكين جديد','transfer':'نقل عامل','temporary_exit':'خروج مؤقت','permanent_exit':'خروج دائم','final_exit':'خروج نهائي','remove':'خروج دائم'},statuses={'pending_supervisors':'بانتظار المشرفين','pending_management':'بانتظار الإدارة','approved':'معتمد','rejected':'مرفوض'})

@app.route('/worker-change-requests/new',methods=['GET','POST'])
@login_required
def new_worker_change_request():
 u=current_user()
 if not can_create_housing(u):abort(403)
 err='';typ=request.form.get('action_type') or ({'delete':'remove'}.get(request.form.get('change_type'),request.form.get('change_type'))) or 'add'
 if request.method=='POST':
  eno=request.form.get('employee_no','').strip();source=request.form.get('source_room','').strip();target=(request.form.get('target_room') or request.form.get('room_no') or '').strip();reason=request.form.get('reason','').strip()
  with closing(conn()) as c:
   worker=None
   if typ in ('transfer','remove','temporary_exit','permanent_exit','final_exit'):
    worker=c.execute('SELECT * FROM workers WHERE employee_no=? AND archived=0',(eno,)).fetchone()
    if not worker:err='العامل غير موجود'
    else: source=worker['room_no']
   target_room=c.execute('SELECT * FROM rooms WHERE room_no=?',(target,)).fetchone() if typ in ('add','transfer') else None
   if not err and typ in ('add','transfer') and not target_room:err='الغرفة الجديدة غير موجودة'
   if not err and typ=='add' and c.execute('SELECT 1 FROM workers WHERE employee_no=? AND archived=0',(eno,)).fetchone():err='الرقم الوظيفي موجود مسبقًا'
   source_sup=supervisor_for_room(source) if source else None; target_sup=supervisor_for_room(target) if target else None
   if not err and typ in ('transfer','remove','temporary_exit','permanent_exit','final_exit') and not source_sup:err='لا يوجد مشرف مسؤول عن الغرفة الحالية'
   if not err and typ in ('add','transfer') and not target_sup:err='لا يوجد مشرف مسؤول عن الغرفة الجديدة'
   if not err:
    no='HA-'+datetime.utcnow().strftime('%Y%m%d%H%M%S%f')[-16:]
    source_status='approved' if not source_sup or source_sup['id']==u['id'] else 'pending'
    target_status='approved' if not target_sup or target_sup['id']==u['id'] else 'pending'
    final_status='pending_management' if source_status=='approved' and target_status=='approved' else 'pending_supervisors'
    vals=(no,typ,worker['id'] if worker else None,eno,request.form.get('iqama_no','').strip() if not worker else worker['iqama_no'],request.form.get('full_name','').strip() if not worker else worker['full_name'],request.form.get('nationality','').strip() if not worker else worker['nationality'],request.form.get('profession','').strip() if not worker else worker['profession'],request.form.get('phone','').strip() if not worker else worker['phone'],source or None,target or None,reason,int(bool(request.form.get('kit_bed'))),int(bool(request.form.get('kit_mattress'))),int(bool(request.form.get('kit_pillow'))),int(bool(request.form.get('kit_sheet'))),int(bool(request.form.get('kit_blanket'))),u['id'],source_sup['id'] if source_sup else None,source_status,target_sup['id'] if target_sup else None,target_status,final_status,now())
    cur=c.execute('''INSERT INTO housing_actions(action_no,action_type,worker_id,employee_no,iqama_no,full_name,nationality,profession,phone,source_room,target_room,reason,kit_bed,kit_mattress,kit_pillow,kit_sheet,kit_blanket,requested_by,source_supervisor_id,source_supervisor_status,target_supervisor_id,target_supervisor_status,final_status,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',vals)
    audit(c,u,'create','housing_action',cur.lastrowid,{'type':typ});c.commit();return redirect(url_for('worker_change_request_detail',qid=cur.lastrowid))
 return page('''<div class="card"><h2>طلب جديد</h2>{% if err %}<p class="err">{{err}}</p>{% endif %}<form method="post"><div class="grid"><div class="field"><label>نوع الطلب</label><select name="action_type"><option value="add">تسكين جديد</option><option value="transfer">نقل عامل</option><option value="temporary_exit">خروج مؤقت</option><option value="permanent_exit">خروج دائم</option><option value="final_exit">خروج نهائي</option></select></div><div class="field"><label>الرقم الوظيفي</label><input name="employee_no" required></div><div class="field"><label>رقم الإقامة</label><input name="iqama_no"></div><div class="field"><label>الاسم</label><input name="full_name"></div><div class="field"><label>الجنسية</label><input name="nationality"></div><div class="field"><label>المهنة</label><input name="profession"></div><div class="field"><label>رقم الجوال</label><input name="phone" inputmode="tel"></div><div class="field"><label>الغرفة الجديدة</label><input name="target_room" value="{{request.args.get('room_no','')}}"></div></div><div class="card"><h3>مستلزمات العامل الجديد فقط</h3><label><input type="checkbox" name="kit_bed" checked> سرير</label> <label><input type="checkbox" name="kit_mattress" checked> مرتبة</label> <label><input type="checkbox" name="kit_pillow" checked> مخدة</label> <label><input type="checkbox" name="kit_sheet" checked> شرشف</label> <label><input type="checkbox" name="kit_blanket" checked> بطانية</label></div><div class="field"><label>سبب الطلب</label><textarea name="reason" required></textarea></div><button class="btn">إرسال الطلب</button></form></div>''','طلب سكن',u,err=err)

@app.route('/worker-change-requests/<int:qid>',methods=['GET','POST'])
@login_required
def worker_change_request_detail(qid):
 u=current_user()
 with closing(conn()) as c:
  q=c.execute('''SELECT q.*,ru.display_name requester,su.display_name source_sup,tu.display_name target_sup,fu.display_name final_decider FROM housing_actions q LEFT JOIN users ru ON ru.id=q.requested_by LEFT JOIN users su ON su.id=q.source_supervisor_id LEFT JOIN users tu ON tu.id=q.target_supervisor_id LEFT JOIN users fu ON fu.id=q.final_decided_by WHERE q.id=?''',(qid,)).fetchone()
  if not q:abort(404)
  visible=is_admin(u) or q['requested_by']==u['id'] or q['source_supervisor_id']==u['id'] or q['target_supervisor_id']==u['id']
  if not visible:abort(403)
  if request.method=='POST':
   decision=request.form.get('decision');reason=request.form.get('decision_reason','').strip()
   if u['role']=='housing_supervisor':
    changed=False
    if q['source_supervisor_id']==u['id'] and q['source_supervisor_status']=='pending':c.execute('UPDATE housing_actions SET source_supervisor_status=?,source_decided_at=? WHERE id=?',(decision,now(),qid));changed=True
    if q['target_supervisor_id']==u['id'] and q['target_supervisor_status']=='pending':c.execute('UPDATE housing_actions SET target_supervisor_status=?,target_decided_at=? WHERE id=?',(decision,now(),qid));changed=True
    if not changed:abort(403)
    fresh=c.execute('SELECT * FROM housing_actions WHERE id=?',(qid,)).fetchone()
    if decision=='rejected' or fresh['source_supervisor_status']=='rejected' or fresh['target_supervisor_status']=='rejected': fs='rejected'
    elif fresh['source_supervisor_status'] in (None,'approved') and fresh['target_supervisor_status'] in (None,'approved'):fs='pending_management'
    else:fs='pending_supervisors'
    c.execute('UPDATE housing_actions SET final_status=?,decision_reason=? WHERE id=?',(fs,reason,qid));audit(c,u,decision,'housing_action',qid)
   elif is_admin(u):
    if q['final_status']!='pending_management':abort(403)
    if decision=='approved':
     if q['action_type']=='add':
      room=c.execute('SELECT * FROM rooms WHERE room_no=?',(q['target_room'],)).fetchone();occupied=c.execute('SELECT COUNT(*) FROM workers WHERE room_no=? AND archived=0',(q['target_room'],)).fetchone()[0]
      if not room or occupied>=room['capacity']:return page('<div class="card"><p class="err">الغرفة غير متاحة أو ممتلئة.</p></div>','خطأ',u),400
      cur=c.execute('''INSERT INTO workers(employee_no,iqama_no,full_name,nationality,profession,phone,zone,room_no,status,archived,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',(q['employee_no'],q['iqama_no'],q['full_name'],q['nationality'],q['profession'],q['phone'],room['zone'],q['target_room'],'active',0,now(),now()))
      c.execute('UPDATE housing_actions SET worker_id=? WHERE id=?',(cur.lastrowid,qid));c.execute('INSERT INTO housing_kit_deliveries(worker_id,bed,mattress,pillow,sheet,blanket,delivered_by,notes,delivered_at) VALUES(?,?,?,?,?,?,?,?,?)',(cur.lastrowid,q['kit_bed'],q['kit_mattress'],q['kit_pillow'],q['kit_sheet'],q['kit_blanket'],q['requested_by'],'تم التسجيل مع التسكين الجديد',now()))
     elif q['action_type']=='transfer':
      room=c.execute('SELECT * FROM rooms WHERE room_no=?',(q['target_room'],)).fetchone();occupied=c.execute('SELECT COUNT(*) FROM workers WHERE room_no=? AND archived=0',(q['target_room'],)).fetchone()[0]
      if not room or occupied>=room['capacity']:return page('<div class="card"><p class="err">الغرفة الجديدة غير متاحة أو ممتلئة.</p></div>','خطأ',u),400
      c.execute('UPDATE workers SET room_no=?,zone=?,status=?,updated_at=? WHERE id=?',(q['target_room'],room['zone'],'active',now(),q['worker_id']))
     elif q['action_type']=='temporary_exit':
      c.execute("UPDATE workers SET status='temporary_exit',updated_at=? WHERE id=?",(now(),q['worker_id']))
     else:
      exit_status='final_exit' if q['action_type']=='final_exit' else 'permanent_exit'
      c.execute('UPDATE workers SET archived=1,status=?,updated_at=? WHERE id=?',(exit_status,now(),q['worker_id']))
    c.execute('UPDATE housing_actions SET final_status=?,final_decided_by=?,decision_reason=?,final_decided_at=? WHERE id=?',(decision,u['id'],reason,now(),qid));audit(c,u,decision,'housing_action',qid)
   else:abort(403)
   c.commit();return redirect(url_for('worker_change_request_detail',qid=qid))
  can_sup=(u['role']=='housing_supervisor' and ((q['source_supervisor_id']==u['id'] and q['source_supervisor_status']=='pending') or (q['target_supervisor_id']==u['id'] and q['target_supervisor_status']=='pending')))
  can_admin=is_admin(u) and q['final_status']=='pending_management'
 return page('''<div class="card"><h2>طلب {{q.action_no}}</h2><p><b>النوع:</b> {{types.get(q.action_type)}}</p><p><b>العامل:</b> {{q.employee_no}} - {{q.full_name}}</p><p><b>من الغرفة:</b> {{q.source_room or '-'}}</p><p><b>إلى الغرفة:</b> {{q.target_room or '-'}}</p><p><b>مقدم الطلب:</b> {{q.requester}}</p><p><b>مشرف الغرفة الحالية:</b> {{q.source_sup or '-'}} — {{q.source_supervisor_status or 'غير مطلوب'}}</p><p><b>مشرف الغرفة الجديدة:</b> {{q.target_sup or '-'}} — {{q.target_supervisor_status or 'غير مطلوب'}}</p><p><b>الاعتماد النهائي:</b> {{q.final_status}}</p><p><b>السبب:</b> {{q.reason}}</p>{% if can_sup or can_admin %}<form method="post"><textarea name="decision_reason" placeholder="ملاحظة القرار"></textarea><br><button class="btn" name="decision" value="approved">موافقة</button> <button class="btn danger" name="decision" value="rejected">رفض</button></form>{% endif %}</div>''','تفاصيل الطلب',u,q=q,can_sup=can_sup,can_admin=can_admin,types={'add':'تسكين جديد','transfer':'نقل عامل','temporary_exit':'خروج مؤقت','permanent_exit':'خروج دائم','final_exit':'خروج نهائي','remove':'خروج دائم'})

@app.get('/housing-kit')
@login_required
def housing_kit_list():
 u=current_user()
 if u['role'] not in ('housing_supervisor','housing_monitor','housing_manager','services_manager'):abort(403)
 cl,args=assigned_clause(u,'r') if u['role']=='housing_supervisor' else ('1=1',[])
 with closing(conn()) as c:rows=c.execute(f'''SELECT w.id,w.employee_no,w.full_name,w.zone,w.room_no,k.bed,k.mattress,k.pillow,k.sheet,k.blanket,k.delivered_at,du.display_name delivered_name FROM workers w JOIN rooms r ON r.room_no=w.room_no LEFT JOIN housing_kit_deliveries k ON k.id=(SELECT id FROM housing_kit_deliveries WHERE worker_id=w.id ORDER BY id DESC LIMIT 1) LEFT JOIN users du ON du.id=k.delivered_by WHERE w.archived=0 AND {cl} ORDER BY CAST(w.room_no AS INTEGER),w.full_name LIMIT 500''',args).fetchall()
 return page('''<h2>تسليم مستلزمات السكن</h2><div class="tbl-wrap"><table class="tbl"><tr><th>العامل</th><th>الغرفة</th><th>سرير</th><th>مرتبة</th><th>مخدة</th><th>شرشف</th><th>بطانية</th><th></th></tr>{% for x in rows %}<tr><td>{{x.employee_no}} - {{x.full_name}}</td><td>{{x.zone}} / {{x.room_no}}</td><td>{{'✓' if x.bed else '—'}}</td><td>{{'✓' if x.mattress else '—'}}</td><td>{{'✓' if x.pillow else '—'}}</td><td>{{'✓' if x.sheet else '—'}}</td><td>{{'✓' if x.blanket else '—'}}</td><td><a href="{{url_for('housing_kit_worker',wid=x.id)}}">تسجيل/تعديل</a></td></tr>{% endfor %}</table></div>''','مستلزمات السكن',u,rows=rows)

@app.route('/workers/<int:wid>/housing-kit',methods=['GET','POST'])
@login_required
def housing_kit_worker(wid):
 u=current_user()
 if u['role'] not in ('housing_supervisor','housing_monitor','housing_manager','services_manager'):abort(403)
 with closing(conn()) as c:
  w=c.execute('SELECT w.*,r.id room_id FROM workers w JOIN rooms r ON r.room_no=w.room_no WHERE w.id=? AND w.archived=0',(wid,)).fetchone()
  if not w:abort(404)
  if u['role']=='housing_supervisor':
   cl,args=assigned_clause(u,'r')
   if not c.execute(f'SELECT 1 FROM rooms r WHERE r.room_no=? AND {cl}',[w['room_no']]+args).fetchone():abort(403)
  if request.method=='POST':
   c.execute('INSERT INTO housing_kit_deliveries(worker_id,bed,mattress,pillow,sheet,blanket,delivered_by,notes,delivered_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?)',(wid,int(bool(request.form.get('bed'))),int(bool(request.form.get('mattress'))),int(bool(request.form.get('pillow'))),int(bool(request.form.get('sheet'))),int(bool(request.form.get('blanket'))),u['id'],request.form.get('notes',''),now(),now()));audit(c,u,'update','housing_kit',wid);c.commit();return redirect(url_for('housing_kit_list'))
  k=c.execute('SELECT * FROM housing_kit_deliveries WHERE worker_id=? ORDER BY id DESC LIMIT 1',(wid,)).fetchone()
 return page('''<div class="card"><h2>مستلزمات {{w.full_name}}</h2><p>{{w.employee_no}} — الغرفة {{w.room_no}}</p><form method="post"><div class="grid"><label><input type="checkbox" name="bed" {% if k and k.bed %}checked{% endif %}> السرير</label><label><input type="checkbox" name="mattress" {% if k and k.mattress %}checked{% endif %}> المرتبة</label><label><input type="checkbox" name="pillow" {% if k and k.pillow %}checked{% endif %}> المخدة</label><label><input type="checkbox" name="sheet" {% if k and k.sheet %}checked{% endif %}> الشرشف</label><label><input type="checkbox" name="blanket" {% if k and k.blanket %}checked{% endif %}> البطانية</label></div><textarea name="notes" placeholder="ملاحظات"></textarea><br><button class="btn">حفظ التسليم</button></form></div>''','مستلزمات العامل',u,w=w,k=k)

@app.get('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
 safe=os.path.basename(filename)
 if safe!=filename or not os.path.isfile(os.path.join(UPLOAD_DIR,safe)):abort(404)
 return send_from_directory(UPLOAD_DIR,safe,max_age=3600)

@app.get('/tickets')
@login_required
def tickets():
 u=current_user();where='';params=[]
 if u['role'] in ('housing_supervisor','housing_monitor','data_entry'):where='WHERE t.reported_by=?';params=[u['id']]
 elif u['role']=='maintenance_supervisor':where='WHERE t.assigned_to=? OR t.assigned_to IS NULL';params=[u['id']]
 with closing(conn()) as c:rows=c.execute(f'''SELECT t.*,u.display_name reporter,a.display_name assignee FROM maintenance_tickets t LEFT JOIN users u ON u.id=t.reported_by LEFT JOIN users a ON a.id=t.assigned_to {where} ORDER BY t.id DESC LIMIT 400''',params).fetchall()
 return page('''<h2>بلاغات الصيانة</h2>{% if not maintenance_only %}<a class="btn" href="{{url_for('new_ticket')}}">بلاغ جديد</a>{% endif %}<table class="tbl"><tr><th>البلاغ</th><th>الموقع</th><th>التصنيف</th><th>الحالة</th><th>المبلغ</th><th></th></tr>{% for x in rows %}<tr><td>{{x.ticket_no}}</td><td>{{x.location_type}} {{x.location_id}}</td><td>{{x.category}}</td><td>{{status_ar.get(x.status,x.status)}}</td><td>{{x.reporter}}</td><td><a href="{{url_for('ticket_detail',tid=x.id)}}">فتح</a></td></tr>{% endfor %}</table>''','بلاغات الصيانة',u,rows=rows,status_ar=STATUS_AR,maintenance_only=is_maintenance_only(u))
@app.route('/tickets/new',methods=['GET','POST'])
@login_required
def new_ticket():
 u=current_user();err=''
 if is_maintenance_only(u):abort(403)
 if request.method=='POST':
  files=request.files.getlist('photos');saved=[]
  try:
   for f in files:
    x=save_upload(f,'ticket');
    if x:saved.append(x)
  except ValueError as e:err=str(e)
  if not saved:err=err or 'يجب إرفاق صورة واحدة على الأقل مع البلاغ'
  if not err:
   no='MNT-'+datetime.utcnow().strftime('%Y%m%d%H%M%S%f')[-16:]
   try:
    with closing(conn()) as c:
     cur=c.execute('INSERT INTO maintenance_tickets(ticket_no,location_type,location_id,zone_name,category,description,priority,status,reported_by,before_photo,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(no,request.form['location_type'],request.form['location_id'],request.form.get('zone_name',''),request.form['category'],request.form.get('description',''),request.form['priority'],'pending_maintenance',u['id'],saved[0],now()))
     for x in saved:c.execute('INSERT INTO ticket_photos(ticket_id,photo_path,photo_kind,uploaded_by,created_at) VALUES(?,?,?,?,?)',(cur.lastrowid,x,'before',u['id'],now()))
     c.execute('INSERT INTO ticket_updates(ticket_id,user_id,action,notes,photo_path,created_at) VALUES(?,?,?,?,?,?)',(cur.lastrowid,u['id'],'created',request.form.get('description',''),saved[0],now()));audit(c,u,'create','maintenance_ticket',cur.lastrowid);c.commit();ticket_id=cur.lastrowid
    return redirect(url_for('ticket_detail',tid=ticket_id))
   except Exception:
    cleanup_uploads(saved)
    err='تعذر إنشاء البلاغ أو ربط الصور. أعد المحاولة، وإذا استمرت المشكلة تواصل مع مدير النظام.'
 selected_location=request.args.get('location_type','room');selected_category=request.args.get('category','')
 categories=['أرضيات','أسقف','جدران','كهرباء','تكييف','سباكة','أثاث','نوافذ','أبواب','أقفال','مقابض','محبس أرضي','محبس ترويش','مروش','مغسلة','تسريب','تصريف','إنارة','أخرى']
 return page('''<div class="card"><h2>بلاغ صيانة جديد</h2>{% if err %}<p class="err">{{err}}</p>{% endif %}<form method="post" enctype="multipart/form-data"><div class="grid"><div class="field"><label>نوع الموقع</label><select name="location_type"><option value="room" {% if selected_location=='room' %}selected{% endif %}>غرفة</option><option value="bathroom" {% if selected_location=='bathroom' %}selected{% endif %}>دورة مياه</option><option value="corridor">ممر</option><option value="kitchen">مطبخ</option><option value="other">أخرى</option></select></div><div class="field"><label>رقم/اسم الموقع</label><input name="location_id" required></div><div class="field"><label>الزون</label><input name="zone_name"></div><div class="field"><label>التصنيف</label><select name="category">{% for x in categories %}<option {% if selected_category==x %}selected{% endif %}>{{x}}</option>{% endfor %}</select></div><div class="field"><label>الأولوية</label><select name="priority"><option value="normal">عادي</option><option value="urgent">عاجل</option><option value="critical">طارئ</option></select></div><div class="field"><label>صور قبل الإصلاح</label><input type="file" name="photos" accept="image/*" capture="environment" multiple required></div></div><div class="field"><label>الوصف</label><textarea name="description" required></textarea></div><button class="btn">إرسال البلاغ للصيانة</button></form></div>''','بلاغ جديد',u,err=err,categories=categories,selected_location=selected_location,selected_category=selected_category)
@app.route('/tickets/<int:tid>',methods=['GET','POST'])
@login_required
def ticket_detail(tid):
 u=current_user()
 with closing(conn()) as c:
  t=c.execute('SELECT * FROM maintenance_tickets WHERE id=?',(tid,)).fetchone()
  if not t:abort(404)
  if u['role'] in ('housing_supervisor','housing_monitor','data_entry') and t['reported_by']!=u['id']:abort(403)
  if u['role']=='maintenance_supervisor' and t['assigned_to'] not in (None,u['id']):abort(403)
  if request.method=='POST':
   action=request.form.get('action');notes=request.form.get('notes','');
   if action=='accept':
    if not (is_maintenance_only(u) or u['role']=='super_admin'):abort(403)
    c.execute('UPDATE maintenance_tickets SET status=?,assigned_to=?,started_at=? WHERE id=?',('accepted',request.form.get('assigned_to') or u['id'],now(),tid))
   elif action=='start':
    if not (is_maintenance_only(u) or u['role']=='super_admin'):abort(403)
    c.execute('UPDATE maintenance_tickets SET status=?,technician_name=?,started_at=? WHERE id=?',('in_progress',request.form.get('technician_name',''),now(),tid))
   elif action=='complete':
    if not (is_maintenance_only(u) or u['role']=='super_admin'):abort(403)
    files=request.files.getlist('after_photos');saved=[]
    try:
     for f in files:
      x=save_upload(f,'ticket_after')
      if x:saved.append(x)
    except ValueError as e:
     cleanup_uploads(saved)
     return page("""<div class='card'><p class='err'>{{err}}</p><a class='btn' href='{{url_for(\"ticket_detail\",tid=tid)}}'>عودة للبلاغ</a></div>""",'خطأ رفع الصور',u,err=str(e),tid=tid),400
    if not saved:return page('<div class="card"><p class="err">صور الإغلاق إلزامية.</p></div>','خطأ',u),400
    for x in saved:c.execute('INSERT INTO ticket_photos(ticket_id,photo_path,photo_kind,uploaded_by,created_at) VALUES(?,?,?,?,?)',(tid,x,'after',u['id'],now()))
    c.execute('UPDATE maintenance_tickets SET status=?,completion_notes=?,after_photo=?,completed_at=? WHERE id=?',('awaiting_reporter',notes,saved[0],now(),tid))
   elif action=='approve_close':
    if t['reported_by']!=u['id'] and not is_admin(u):abort(403)
    c.execute('UPDATE maintenance_tickets SET status=?,verification_by=?,verified_at=?,closed_at=? WHERE id=?',('closed',u['id'],now(),now(),tid))
   elif action=='return':
    if t['reported_by']!=u['id'] and not is_admin(u):abort(403)
    c.execute('UPDATE maintenance_tickets SET status=?,completion_notes=? WHERE id=?',('returned',notes,tid))
   else:abort(400)
   c.execute('INSERT INTO ticket_updates(ticket_id,user_id,action,notes,created_at) VALUES(?,?,?,?,?)',(tid,u['id'],action,notes,now()));audit(c,u,action,'maintenance_ticket',tid);c.commit();return redirect(url_for('ticket_detail',tid=tid))
  maintenance_users=c.execute("SELECT id,display_name FROM users WHERE active=1 AND role IN ('maintenance_manager','maintenance_supervisor') ORDER BY role,display_name").fetchall()
  updates=c.execute('''SELECT x.*,u.display_name FROM ticket_updates x LEFT JOIN users u ON u.id=x.user_id WHERE x.ticket_id=? ORDER BY x.id DESC''',(tid,)).fetchall();photos=c.execute('SELECT * FROM ticket_photos WHERE ticket_id=? ORDER BY id',(tid,)).fetchall()
 before=[x for x in photos if x['photo_kind'] in ('report','before')];after=[x for x in photos if x['photo_kind']=='after']
 return page('''<div class="card"><h2>بلاغ {{ticket.ticket_no}}</h2><div class="photo-stage"><h3>صور قبل الإصلاح</h3><div class="grid">{% for p in before %}<a href="{{url_for('uploaded_file',filename=p.photo_path)}}" target="_blank"><img src="{{url_for('uploaded_file',filename=p.photo_path)}}"></a>{% endfor %}</div><div class="card"><p><b>الموقع:</b> {{ticket.location_type}} {{ticket.location_id}}</p><p><b>التصنيف:</b> {{ticket.category}}</p><p><b>الوصف:</b> {{ticket.description}}</p><p><b>الحالة:</b> {{status_ar.get(t.status,t.status)}}</p></div><h3>صور بعد الإصلاح</h3><div class="grid">{% for p in after %}<a href="{{url_for('uploaded_file',filename=p.photo_path)}}" target="_blank"><img src="{{url_for('uploaded_file',filename=p.photo_path)}}"></a>{% else %}<p class="muted">لم تُرفع صور الإغلاق بعد.</p>{% endfor %}</div></div>
 {% if maintenance_only and ticket.status in ('new','pending_maintenance','returned') %}<form method="post"><select name="assigned_to">{% for m in maintenance_users %}<option value="{{m.id}}">{{m.display_name}}</option>{% endfor %}</select><button class="btn" name="action" value="accept">قبول البلاغ</button></form>{% endif %}
 {% if maintenance_only and ticket.status in ('accepted','returned') %}<form method="post"><input name="technician_name" placeholder="اسم الفني"><button class="btn" name="action" value="start">بدء التنفيذ</button></form>{% endif %}
 {% if maintenance_only and ticket.status=='in_progress' %}<form method="post" enctype="multipart/form-data"><textarea name="notes" placeholder="ملاحظات التنفيذ" required></textarea><input type="file" name="after_photos" accept="image/*" multiple required><button class="btn" name="action" value="complete">تم التنفيذ وإرسال للمبلغ</button></form>{% endif %}
 {% if can_verify and ticket.status=='awaiting_reporter' %}<form method="post"><textarea name="notes" placeholder="ملاحظة التحقق"></textarea><button class="btn" name="action" value="approve_close">اعتماد الإغلاق</button> <button class="btn danger" name="action" value="return">إعادة للصيانة</button></form>{% endif %}</div><div class="card"><h3>سجل البلاغ</h3><table class="tbl">{% for x in updates %}<tr><td>{{x.created_at}}</td><td>{{x.display_name}}</td><td>{{x.action}}</td><td>{{x.notes}}</td></tr>{% endfor %}</table></div>''','تفاصيل البلاغ',u,ticket=t,updates=updates,status_ar=STATUS_AR,maintenance_only=(is_maintenance_only(u) or u['role']=='super_admin'),maintenance_users=maintenance_users,before=before,after=after,can_verify=(t['reported_by']==u['id'] or is_admin(u)))

@app.get('/occupancy-management')
@login_required
def occupancy_management():
 u=current_user();cl,args=assigned_clause(u,'r');zone=request.args.get('zone','').strip();q=request.args.get('q','').strip();state=request.args.get('state','').strip()
 sql=f'''SELECT r.*,COUNT(w.id) occupied,(r.capacity-COUNT(w.id)) free_beds FROM rooms r LEFT JOIN workers w ON w.room_no=r.room_no AND w.archived=0 WHERE {cl}''';params=list(args)
 if zone:sql+=' AND r.zone=?';params.append(zone)
 if q:sql+=' AND r.room_no LIKE ?';params.append(f'%{q}%')
 sql+=' GROUP BY r.id ORDER BY CAST(r.room_no AS INTEGER),r.room_no'
 with closing(conn()) as c:
  rows=c.execute(sql,params).fetchall();zones=[x[0] for x in c.execute('SELECT DISTINCT zone FROM rooms ORDER BY zone').fetchall()]
 def room_state(x):
  if (x['status'] or '') in ('closed','out_of_service') or (x['usage_type'] or '')=='closed':return 'closed'
  if (x['usage_type'] or '')=='maintenance':return 'maintenance'
  if x['occupied']>x['capacity']:return 'crowded'
  if x['occupied']==0 and (x['usage_type'] or 'residential')=='residential':return 'vacant'
  return 'occupied'
 decorated=[dict(x)|{'room_state':room_state(x)} for x in rows]
 if state:decorated=[x for x in decorated if x['room_state']==state]
 counts={k:sum(1 for x in decorated if x['room_state']==k) for k in ('occupied','vacant','crowded','closed','maintenance')}
 return page('''<h2>{{t.occupancy}}</h2><div class="cards"><div class="card"><b>مشغولة</b><div class="num">{{counts.occupied}}</div></div><div class="card"><b>شاغرة</b><div class="num">{{counts.vacant}}</div></div><div class="card"><b>متكدسة</b><div class="num">{{counts.crowded}}</div></div><div class="card"><b>مغلقة</b><div class="num">{{counts.closed}}</div></div><div class="card"><b>صيانة</b><div class="num">{{counts.maintenance}}</div></div></div><form class="card"><div class="grid"><input name="q" value="{{q}}" placeholder="رقم الغرفة"><select name="zone"><option value="">كل الزونات</option>{% for z in zones %}<option {% if zone==z|string %}selected{% endif %}>{{z}}</option>{% endfor %}</select><select name="state"><option value="">كل الحالات</option><option value="occupied">مشغولة</option><option value="vacant">شاغرة</option><option value="crowded">متكدسة</option><option value="closed">مغلقة</option><option value="maintenance">صيانة</option></select></div><button class="btn">بحث</button></form><div class="tbl-wrap"><table class="tbl"><tr><th>الزون</th><th>الغرفة</th><th>السعة</th><th>المقيمون</th><th>الحالة</th><th></th></tr>{% for x in rows %}<tr><td>{{x.zone}}</td><td>{{x.room_no}}</td><td>{{x.capacity}}</td><td>{{x.occupied}}</td><td>{{x.room_state}}</td><td><a href="{{url_for('room_detail',room_no=x.room_no)}}">فتح</a></td></tr>{% endfor %}</table></div>''',tr('occupancy'),u,rows=decorated,zones=zones,q=q,zone=zone,state=state,counts=counts)

@app.get('/occupancy-map')
@login_required
def occupancy_map():
 # Phase 5: the occupancy map was retired; sectors are the operational replacement.
 return redirect(url_for('sector_dashboard'))

@app.get('/sector')
@login_required
def sector_dashboard():
 u=current_user();cl,args=assigned_clause(u,'r')
 with closing(conn()) as c:
  rooms=c.execute(f'''SELECT r.*,COUNT(w.id) occupied FROM rooms r LEFT JOIN workers w ON w.room_no=r.room_no AND w.archived=0 WHERE {cl} GROUP BY r.id ORDER BY CAST(r.zone AS INTEGER),CAST(r.room_no AS INTEGER)''',args).fetchall()
  zones_for_user=assigned_zones(c,u)
  if u['role']=='housing_supervisor' and zones_for_user:
   marks=','.join('?' for _ in zones_for_user); assets_where=f'WHERE CAST(b.zone_name AS TEXT) IN ({marks})'; assets_params=zones_for_user
  elif u['role']=='housing_supervisor':
   assets_where='WHERE 1=0'; assets_params=[]
  else:
   assets_where=''; assets_params=[]
  assets=c.execute(f'''SELECT a.*,b.name complex_name,b.zone_name,(SELECT MAX(created_at) FROM bathroom_asset_inspections i WHERE i.asset_id=a.id) last_inspection FROM bathroom_assets a JOIN bathroom_complexes b ON b.id=a.complex_id {assets_where} ORDER BY CAST(b.zone_name AS INTEGER),a.asset_type,a.asset_no''',assets_params).fetchall()
  today=date.today().isoformat()
  inspected_today=c.execute("SELECT COUNT(DISTINCT location_id) FROM inspections WHERE inspection_type='room' AND substr(created_at,1,10)=?"+(" AND inspector_id=?" if u['role']=='housing_supervisor' else ''),([today,u['id']] if u['role']=='housing_supervisor' else [today])).fetchone()[0]
 return page('''<h2>القطاع والجولة</h2><div class="cards"><div class="card"><b>إجمالي الغرف</b><div class="num">{{rooms|length}}</div></div><div class="card"><b>المكتمل اليوم</b><div class="num">{{inspected_today}}</div></div><div class="card"><b>المتبقي اليوم</b><div class="num">{{[rooms|length-inspected_today,0]|max}}</div></div><div class="card"><b>دورات المياه</b><div class="num">{{assets|selectattr('asset_type','equalto','toilet')|list|length}}</div></div><div class="card"><b>المغاسل</b><div class="num">{{assets|selectattr('asset_type','equalto','basin')|list|length}}</div></div></div><div class="card"><h3>غرف القطاع</h3><p class="muted">تبدأ الجولة بالغرف، ثم تنتقل إلى دورات المياه والمغاسل المرقمة.</p><div class="room-map">{% for x in rooms %}<a class="room-tile {{'yellow' if x.occupied==0 else 'green'}}" href="{{url_for('new_inspection',room_no=x.room_no)}}"><b>{{x.room_no}}</b><span>{{x.zone}} · {{x.occupied}}/{{x.capacity}}</span></a>{% endfor %}</div></div><div class="card"><h3>دورات المياه والمغاسل</h3><div class="tbl-wrap"><table class="tbl"><tr><th>المجمع</th><th>النوع</th><th>الرقم</th><th>آخر فحص</th><th></th></tr>{% for x in assets %}<tr><td>{{x.complex_name}}</td><td>{{'دورة مياه' if x.asset_type=='toilet' else 'مغسلة'}}</td><td>{{x.asset_no}}</td><td>{{x.last_inspection or 'لم يفحص'}}</td><td>{% if u.role=='housing_supervisor' %}<a href="{{url_for('bathroom_asset_inspection',asset_id=x.id)}}">ابدأ الفحص</a>{% else %}<span class="muted">عرض فقط</span>{% endif %}</td></tr>{% endfor %}</table></div></div>''','القطاع والجولة',u,rooms=rooms,assets=assets,inspected_today=inspected_today,u=u)

@app.route('/bathroom-assets/<int:asset_id>/inspect',methods=['GET','POST'])
@login_required
def bathroom_asset_inspection(asset_id):
 u=current_user()
 if u['role']!='housing_supervisor':abort(403)
 err=''
 with closing(conn()) as c:
  zones_for_user=assigned_zones(c,u)
  if not zones_for_user:abort(404)
  marks=','.join('?' for _ in zones_for_user)
  asset=c.execute(f'''SELECT a.*,b.name complex_name,b.zone_name FROM bathroom_assets a JOIN bathroom_complexes b ON b.id=a.complex_id WHERE a.id=? AND CAST(b.zone_name AS TEXT) IN ({marks})''',[asset_id]+zones_for_user).fetchone()
  if not asset:abort(404)
  if request.method=='POST':
   condition=request.form.get('condition_status','ممتاز');notes=request.form.get('notes','').strip();mreq=int(bool(request.form.get('maintenance_required')));photo=None;ticket_id=None
   try:photo=save_upload(request.files.get('photo'),'bathroom')
   except ValueError as e:err=str(e)
   if not err:
    if mreq:
     desc=request.form.get('maintenance_description','').strip() or notes
     if not desc:err='اكتب وصف العطل'
     else:
      ticket_no='MNT-'+datetime.utcnow().strftime('%Y%m%d%H%M%S%f')[-16:]
      t=c.execute('INSERT INTO maintenance_tickets(ticket_no,location_type,location_id,zone_name,category,description,priority,status,reported_by,before_photo,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(ticket_no,asset['asset_type'],f"{asset['complex_name']} / {asset['asset_no']}",asset['zone_name'],'دورات المياه والمغاسل',desc,'normal','new',u['id'],photo,now()));ticket_id=t.lastrowid
    if not err:
     cur=c.execute('INSERT INTO bathroom_asset_inspections(asset_id,inspector_id,condition_status,notes,photo_path,maintenance_required,maintenance_ticket_id,created_at) VALUES(?,?,?,?,?,?,?,?)',(asset_id,u['id'],condition,notes,photo,mreq,ticket_id,now()));audit(c,u,'create','bathroom_asset_inspection',cur.lastrowid,{'asset_id':asset_id});c.commit();return redirect(url_for('sector_dashboard'))
  history=c.execute('''SELECT i.*,u.display_name FROM bathroom_asset_inspections i LEFT JOIN users u ON u.id=i.inspector_id WHERE i.asset_id=? ORDER BY i.id DESC LIMIT 30''',(asset_id,)).fetchall()
 return page('''<div class="card"><h2>{{asset.complex_name}}</h2><h3>{{'دورة مياه' if asset.asset_type=='toilet' else 'مغسلة'}} رقم {{asset.asset_no}}</h3>{% if err %}<p class="err">{{err}}</p>{% endif %}<form method="post" enctype="multipart/form-data"><div class="grid"><div class="field"><label>الحالة</label><select name="condition_status"><option>ممتاز</option><option>يحتاج متابعة</option><option>مغلق بسبب مستخدم</option><option>خارج الخدمة</option></select></div><div class="field"><label>صورة</label><input type="file" name="photo" accept="image/*" capture="environment"></div></div><div class="field"><label>الملاحظات</label><textarea name="notes"></textarea></div><label><input type="checkbox" name="maintenance_required" value="1"> إنشاء بلاغ صيانة</label><div class="field"><label>وصف العطل</label><textarea name="maintenance_description"></textarea></div><button class="btn">حفظ الفحص</button></form></div><div class="card"><h3>السجل</h3><table class="tbl"><tr><th>التاريخ</th><th>الحالة</th><th>المشرف</th><th>الملاحظات</th></tr>{% for x in history %}<tr><td>{{x.created_at}}</td><td>{{x.condition_status}}</td><td>{{x.display_name}}</td><td>{{x.notes}}</td></tr>{% endfor %}</table></div>''','فحص المرفق',u,asset=asset,history=history,err=err)

@app.route('/workers/<int:wid>/absence-report',methods=['GET','POST'])
@login_required
def new_absence_report(wid):
 u=current_user()
 if u['role']!='housing_supervisor':abort(403)
 with closing(conn()) as c:
  w=c.execute('''SELECT w.* FROM workers w JOIN rooms r ON r.room_no=w.room_no WHERE w.id=? AND w.archived=0 AND r.supervisor_id=?''',(wid,u['id'])).fetchone()
  if not w:abort(404)
  if request.method=='POST':
   cur=c.execute('INSERT INTO absence_reports(worker_id,room_no,reported_by,absence_duration,notes,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?)',(wid,w['room_no'],u['id'],request.form.get('absence_duration',''),request.form.get('notes',''),'new',now(),now()));c.execute('INSERT INTO worker_events(worker_id,event_type,description,user_id,related_id,created_at) VALUES(?,?,?,?,?,?)',(wid,'absence_report',request.form.get('notes',''),u['id'],cur.lastrowid,now()));audit(c,u,'create','absence_report',cur.lastrowid);c.commit();return redirect(url_for('absence_reports_list'))
 return page('''<div class="card"><h2>بلاغ عامل غير متواجد</h2><p><b>{{w.full_name}}</b> — {{w.employee_no}} — الغرفة {{w.room_no}}</p><form method="post"><div class="field"><label>مدة عدم التواجد</label><input name="absence_duration" placeholder="مثال: شهر" required></div><div class="field"><label>الملاحظات</label><textarea name="notes" required></textarea></div><button class="btn">رفع البلاغ لمدير السكن</button></form></div>''','بلاغ عدم تواجد',u,w=w)

@app.route('/absence-reports',methods=['GET','POST'])
@login_required
def absence_reports_list():
 u=current_user()
 with closing(conn()) as c:
  if request.method=='POST':
   if not is_admin(u):abort(403)
   rid=int(request.form['report_id']);status=request.form['status'];notes=request.form.get('manager_notes','')
   c.execute('UPDATE absence_reports SET status=?,manager_id=?,manager_notes=?,updated_at=? WHERE id=?',(status,u['id'],notes,now(),rid));audit(c,u,'decision','absence_report',rid,{'status':status});c.commit();return redirect(url_for('absence_reports_list'))
  where='' if is_admin(u) else 'WHERE a.reported_by=?';params=[] if is_admin(u) else [u['id']]
  rows=c.execute(f'''SELECT a.*,w.employee_no,w.full_name,u.display_name reporter FROM absence_reports a JOIN workers w ON w.id=a.worker_id LEFT JOIN users u ON u.id=a.reported_by {where} ORDER BY a.id DESC''',params).fetchall()
 return page('''<h2>بلاغات العامل غير المتواجد</h2><p class="muted">افتح البلاغ لمتابعته واتخاذ الإجراء وتسجيل الملاحظات.</p>{% if role=='housing_supervisor' %}<a class="btn" href="{{url_for('new_absence_by_employee')}}">رفع بلاغ جديد بالرقم الوظيفي</a>{% endif %}<div class="tbl-wrap"><table class="tbl"><tr><th>العامل</th><th>الغرفة</th><th>المدة</th><th>المشرف</th><th>الحالة</th><th>الملاحظات</th><th>الإجراء</th></tr>{% for x in rows %}<tr><td>{{x.employee_no}} - {{x.full_name}}</td><td>{{x.room_no}}</td><td>{{x.absence_duration}}</td><td>{{x.reporter}}</td><td><span class="badge">{{x.status}}</span></td><td>{{x.notes}}</td><td><a class="btn" href="{{url_for('absence_report_detail',rid=x.id)}}">فتح البلاغ</a></td></tr>{% else %}<tr><td colspan="7" class="muted">لا توجد بلاغات عدم تواجد حالياً.</td></tr>{% endfor %}</table></div>''','بلاغات عدم التواجد',u,rows=rows,admin=is_admin(u),role=u['role'])

@app.route('/absence-reports/new',methods=['GET','POST'])
@login_required
def new_absence_by_employee():
 u=current_user()
 if u['role']!='housing_supervisor':abort(403)
 err='';worker=None;eno=(request.form.get('employee_no') or request.args.get('employee_no') or '').strip()
 with closing(conn()) as c:
  if eno:
   worker=c.execute('''SELECT w.*,r.sector_name FROM workers w JOIN rooms r ON r.room_no=w.room_no WHERE w.employee_no=? AND w.archived=0 AND r.supervisor_id=?''',(eno,u['id'])).fetchone()
   if not worker:err='لم يتم العثور على العامل ضمن نطاق الغرف المسندة لك.'
  if request.method=='POST' and request.form.get('action')=='submit' and worker:
   duration=request.form.get('absence_duration','').strip();notes=request.form.get('notes','').strip()
   if not duration:err='مدة عدم التواجد مطلوبة.'
   elif not notes:err='الملاحظات مطلوبة.'
   else:
    cur=c.execute('INSERT INTO absence_reports(worker_id,room_no,reported_by,absence_duration,notes,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?)',(worker['id'],worker['room_no'],u['id'],duration,notes,'new',now(),now()))
    c.execute('INSERT INTO worker_events(worker_id,event_type,description,user_id,related_id,created_at) VALUES(?,?,?,?,?,?)',(worker['id'],'absence_report',notes,u['id'],cur.lastrowid,now()))
    audit(c,u,'create','absence_report',cur.lastrowid,{'employee_no':eno,'duration':duration});c.commit();return redirect(url_for('absence_reports_list'))
 return page('''<div class="card"><h2>بلاغ عامل غير متواجد</h2>{% if err %}<p class="err">{{err}}</p>{% endif %}<form method="post"><div class="grid"><div class="field"><label>الرقم الوظيفي</label><input name="employee_no" value="{{eno}}" required></div><div class="field"><label>&nbsp;</label><button class="btn" name="action" value="lookup">استدعاء بيانات العامل</button></div></div>{% if worker %}<div class="card"><div class="grid"><p><b>الاسم:</b> {{worker.full_name}}</p><p><b>الإقامة:</b> {{worker.iqama_no or '-'}}</p><p><b>الجنسية:</b> {{worker.nationality}}</p><p><b>المهنة:</b> {{worker.profession}}</p><p><b>الزون:</b> {{worker.zone}}</p><p><b>الغرفة:</b> {{worker.room_no}}</p><p><b>القطاع:</b> {{worker.sector_name or '-'}}</p><p><b>الحالة:</b> {{worker.status}}</p></div></div><div class="field"><label>مدة عدم التواجد</label><input name="absence_duration" placeholder="مثال: شهر" required></div><div class="field"><label>الملاحظات</label><textarea name="notes" required></textarea></div><button class="btn" name="action" value="submit">رفع البلاغ لمدير السكن</button>{% endif %}</form></div>''','بلاغ عدم تواجد',u,eno=eno,worker=worker,err=err)

@app.get('/available-beds')
@login_required
def available_beds():
 u=current_user();cl,args=assigned_clause(u,'r')
 zone=request.args.get('zone','').strip();q=request.args.get('q','').strip()
 with closing(conn()) as c:
  sql=f'''SELECT r.room_no,r.zone,r.capacity,r.sector_name,r.supervisor_id,u.display_name supervisor_name,COUNT(w.id) occupied,(r.capacity-COUNT(w.id)) free_beds FROM rooms r LEFT JOIN workers w ON w.room_no=r.room_no AND w.archived=0 LEFT JOIN users u ON u.id=r.supervisor_id WHERE {cl} AND COALESCE(r.usage_type,'residential')='residential' AND COALESCE(r.status,'active')='active' '''
  params=list(args)
  if zone:sql+=' AND CAST(r.zone AS TEXT)=?';params.append(zone)
  if q:sql+=' AND (r.room_no LIKE ? OR r.sector_name LIKE ? OR u.display_name LIKE ?)';like=f'%{q}%';params += [like]*3
  sql+=' GROUP BY r.id HAVING COUNT(w.id)<COALESCE(r.capacity,0) ORDER BY CAST(r.zone AS INTEGER),CAST(r.room_no AS INTEGER)'
  rows=c.execute(sql,params).fetchall();zones=[x['zone'] for x in c.execute(f'''SELECT DISTINCT r.zone FROM rooms r WHERE {cl} AND COALESCE(r.usage_type,'residential')='residential' ORDER BY CAST(r.zone AS INTEGER)''',args).fetchall()]
  total=sum(max(x['free_beds'],0) for x in rows)
 return page('''<h2>الغرف المتاحة للتسكين</h2><p class="muted">تظهر لك الغرف الواقعة ضمن صلاحياتك فقط، بينما أصحاب الصلاحية العامة يرون جميع الغرف.</p><div class="cards"><div class="card"><div>إجمالي الأسرة المتاحة في النتائج</div><div class="num">{{total}}</div></div><div class="card"><div>عدد الغرف المتاحة</div><div class="num">{{rows|length}}</div></div></div><form class="card" method="get"><div class="grid"><div class="field"><label>بحث</label><input name="q" value="{{q}}" placeholder="رقم الغرفة أو القطاع أو المشرف"></div><div class="field"><label>الزون</label><select name="zone"><option value="">كل الزونات</option>{% for z in zones %}<option value="{{z}}" {% if zone|string==z|string %}selected{% endif %}>زون {{z}}</option>{% endfor %}</select></div></div><button class="btn">تصفية</button> <a class="btn btn2" href="{{url_for('available_beds')}}">مسح</a></form><div class="tbl-wrap"><table class="tbl"><tr><th>الزون</th><th>الغرفة</th><th>القطاع</th><th>المشرف</th><th>السعة</th><th>الموجود</th><th>الشاغر</th><th>الإجراء</th></tr>{% for x in rows %}<tr><td>{{x.zone}}</td><td><b>{{x.room_no}}</b></td><td>{{x.sector_name or '-'}}</td><td>{{x.supervisor_name or 'إدارة السكن'}}</td><td>{{x.capacity}}</td><td>{{x.occupied}}</td><td><span class="badge">{{x.free_beds}}</span></td><td><a class="btn" href="{{url_for('new_worker_change_request',room_no=x.room_no)}}">طلب تسكين</a> <a class="btn btn2" href="{{url_for('room_detail',room_no=x.room_no)}}">فتح الغرفة</a></td></tr>{% else %}<tr><td colspan="8" class="muted">لا توجد غرف متاحة ضمن نطاقك الحالي.</td></tr>{% endfor %}</table></div>''','الأسرة الشاغرة',u,rows=rows,total=total,zones=zones,zone=zone,q=q)

@app.route('/absence-reports/<int:rid>',methods=['GET','POST'])
@login_required
def absence_report_detail(rid):
 u=current_user()
 with closing(conn()) as c:
  a=c.execute('''SELECT a.*,w.employee_no,w.full_name,w.zone,w.profession,ru.display_name reporter,mu.display_name manager FROM absence_reports a JOIN workers w ON w.id=a.worker_id LEFT JOIN users ru ON ru.id=a.reported_by LEFT JOIN users mu ON mu.id=a.manager_id WHERE a.id=?''',(rid,)).fetchone()
  if not a:abort(404)
  if not is_admin(u) and a['reported_by']!=u['id']:abort(403)
  if request.method=='POST':
   if not is_admin(u):abort(403)
   status=request.form.get('status');notes=request.form.get('manager_notes','').strip()
   if status not in ('follow_up','closed','rejected'):abort(400)
   if status=='rejected' and not notes:return page("""<div class='card'><p class='err'>سبب الرفض مطلوب.</p><a class='btn' href='{{url_for(\"absence_report_detail\",rid=rid)}}'>رجوع</a></div>""",'خطأ',u,rid=rid),400
   c.execute('UPDATE absence_reports SET status=?,manager_id=?,manager_notes=?,updated_at=? WHERE id=?',(status,u['id'],notes,now(),rid));audit(c,u,'absence_'+status,'absence_report',rid,{'notes':notes});c.commit();return redirect(url_for('absence_report_detail',rid=rid))
  history=c.execute('''SELECT l.*,u.display_name FROM audit_logs l LEFT JOIN users u ON u.id=l.user_id WHERE l.entity_type='absence_report' AND l.entity_id=? ORDER BY l.id DESC''',(rid,)).fetchall()
 return page('''<div class="card"><h2>بلاغ عدم تواجد رقم {{a.id}}</h2><div class="grid"><p><b>العامل:</b> {{a.employee_no}} - {{a.full_name}}</p><p><b>الغرفة:</b> {{a.room_no}}</p><p><b>الزون:</b> {{a.zone}}</p><p><b>المهنة:</b> {{a.profession or '-'}}</p><p><b>مقدم البلاغ:</b> {{a.reporter}}</p><p><b>المدة:</b> {{a.absence_duration}}</p><p><b>الحالة:</b> <span class="badge">{{a.status}}</span></p><p><b>تاريخ البلاغ:</b> {{a.created_at}}</p></div><p><b>ملاحظات مقدم البلاغ:</b> {{a.notes}}</p>{% if a.manager_notes %}<p><b>ملاحظات الإدارة:</b> {{a.manager_notes}}</p>{% endif %}{% if admin and a.status not in ('closed','rejected') %}<form method="post"><div class="field"><label>ملاحظة الإجراء</label><textarea name="manager_notes"></textarea></div><button class="btn" name="status" value="follow_up">تحت المتابعة</button> <button class="btn" name="status" value="closed">إغلاق البلاغ</button> <button class="btn danger" name="status" value="rejected">رفض البلاغ</button></form>{% endif %}</div><div class="card"><h3>سجل الإجراءات</h3><div class="timeline">{% for x in history %}<div class="timeline-item"><b>{{x.action}}</b><div>{{x.created_at}} · {{x.display_name or x.username or '-'}}</div><small>{{x.details_json or '-'}}</small></div>{% else %}<p class="muted">لا توجد إجراءات إضافية حتى الآن.</p>{% endfor %}</div></div>''','تفاصيل بلاغ عدم التواجد',u,a=a,history=history,admin=is_admin(u))

@app.get('/exports/workers.xlsx')
@login_required
def export_workers_excel():
 u=current_user()
 if not can_export(u):abort(403)
 from openpyxl import Workbook
 from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
 from openpyxl.utils import get_column_letter
 q,zone,status=export_filters()
 with closing(conn()) as c:rows=worker_export_rows(c,q,zone,status)
 wb=Workbook();ws=wb.active;ws.title='بيانات العمالة';ws.sheet_view.rightToLeft=True
 headers=['م','الرقم الوظيفي','رقم الإقامة','الاسم الكامل','الجنسية','المهنة','رقم الجوال','الزون','رقم الغرفة','القطاع','سعة الغرفة','استخدام الغرفة','حالة الغرفة','مشرف الغرفة','حالة العامل','تاريخ الإضافة','آخر تحديث']
 ws.append(headers)
 for i,r in enumerate(rows,1):ws.append([i,r['employee_no'],r['iqama_no'],r['full_name'],r['nationality'],r['profession'],r['phone'],r['zone'],r['room_no'],r['sector_name'],r['capacity'],ROOM_USAGE_AR.get(r['usage_type'],r['usage_type']),r['room_status'],r['supervisor_name'],r['status'],r['created_at'],r['updated_at']])
 fill=PatternFill('solid',fgColor='123B32');font=Font(color='FFFFFF',bold=True);thin=Side(style='thin',color='D9E1DE')
 for cell in ws[1]:cell.fill=fill;cell.font=font;cell.alignment=Alignment(horizontal='center',vertical='center');cell.border=Border(bottom=thin)
 for row in ws.iter_rows(min_row=2):
  for cell in row:cell.alignment=Alignment(horizontal='right',vertical='center');cell.border=Border(bottom=thin)
 widths=[7,16,18,34,16,24,17,9,13,18,12,18,14,24,17,20,20]
 for idx,width in enumerate(widths,1):ws.column_dimensions[get_column_letter(idx)].width=width
 ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions
 bio=io.BytesIO();wb.save(bio);bio.seek(0)
 with closing(conn()) as c:audit(c,u,'export','workers_excel',None,{'count':len(rows),'q':q,'zone':zone,'status':status});c.commit()
 return send_file(bio,as_attachment=True,download_name=f'workers_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get('/exports/workers.pdf')
@login_required
def export_workers_pdf():
 u=current_user()
 if not can_export(u):abort(403)
 from reportlab.lib.pagesizes import A3,landscape
 from reportlab.pdfgen import canvas
 from reportlab.pdfbase import pdfmetrics
 from reportlab.pdfbase.ttfonts import TTFont
 from reportlab.lib import colors
 import re
 q,zone,status=export_filters()
 with closing(conn()) as c:rows=worker_export_rows(c,q,zone,status)
 pdfmetrics.registerFont(TTFont('Arabic','/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
 try:
  import arabic_reshaper
  def shape(v):
   t='' if v is None else str(v)
   if not re.search(r'[\u0600-\u06FF]',t): return t
   reshaped=arabic_reshaper.reshape(t)
   try:
    from bidi.algorithm import get_display
    return get_display(reshaped)
   except Exception:
    return reshaped[::-1]
 except Exception:
  def shape(v):return '' if v is None else str(v)
 bio=io.BytesIO();page_size=landscape(A3);w,h=page_size;cvs=canvas.Canvas(bio,pagesize=page_size)
 cols=[('م',18),('الرقم الوظيفي',64),('الإقامة',76),('الاسم',170),('الجنسية',70),('المهنة',110),('الجوال',78),('الزون',38),('الغرفة',48),('القطاع',65),('المشرف',105),('الحالة',65)]
 margin=18;row_h=12;header_h=22
 def draw_header(page_no):
  cvs.setFont('Arabic',12);cvs.drawCentredString(w/2,h-18,shape('تقرير جميع بيانات العمالة - سكن ولي العهد'))
  cvs.setFont('Arabic',6);cvs.drawRightString(w-margin,h-31,shape(f'عدد السجلات: {len(rows)} | تاريخ الاستخراج: {datetime.now().strftime("%Y-%m-%d %H:%M")} | صفحة {page_no}'))
  y=h-48;x=margin;cvs.setFillColor(colors.HexColor('#123B32'));cvs.rect(margin,y-header_h,w-2*margin,header_h,fill=1,stroke=0);cvs.setFillColor(colors.white);cvs.setFont('Arabic',6)
  for title,width in cols:
   cvs.drawCentredString(x+width/2,y-14,shape(title));x+=width
  cvs.setFillColor(colors.black);return y-header_h
 page=1;y=draw_header(page);cvs.setFont('Arabic',5.2)
 for i,r in enumerate(rows,1):
  if y-row_h < 18:
   cvs.showPage();page+=1;y=draw_header(page);cvs.setFont('Arabic',5.2)
  vals=[i,r['employee_no'],r['iqama_no'] or '-',r['full_name'],r['nationality'],r['profession'],r['phone'] or '-',r['zone'],r['room_no'],r['sector_name'] or '-',r['supervisor_name'] or '-',r['status']]
  if i%2==0:
   cvs.setFillColor(colors.HexColor('#F2F6F4'));cvs.rect(margin,y-row_h,w-2*margin,row_h,fill=1,stroke=0);cvs.setFillColor(colors.black)
  x=margin
  for val,(_,width) in zip(vals,cols):
   txt=shape(val)
   maxchars=max(3,int(width/3.2));txt=txt if len(txt)<=maxchars else txt[:maxchars-1]+'…'
   cvs.drawCentredString(x+width/2,y-8,txt);x+=width
  cvs.setStrokeColor(colors.HexColor('#D9E1DE'));cvs.line(margin,y-row_h,w-margin,y-row_h);y-=row_h
 cvs.save();bio.seek(0)
 with closing(conn()) as c:audit(c,u,'export','workers_pdf',None,{'count':len(rows),'q':q,'zone':zone,'status':status});c.commit()
 return send_file(bio,as_attachment=True,download_name=f'workers_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',mimetype='application/pdf')

@app.get('/reports')
@login_required
def reports_center():
 u=current_user()
 if not is_admin(u):abort(403)
 with closing(conn()) as c:
  k=c.execute('''SELECT (SELECT COUNT(*) FROM workers WHERE archived=0) workers,(SELECT COUNT(*) FROM rooms WHERE usage_type='residential') residential_rooms,(SELECT COALESCE(SUM(capacity),0) FROM rooms WHERE usage_type='residential') capacity,(SELECT COUNT(*) FROM rooms r WHERE r.usage_type='residential' AND NOT EXISTS(SELECT 1 FROM workers w WHERE w.room_no=r.room_no AND w.archived=0)) vacant_rooms,(SELECT COUNT(*) FROM absence_reports WHERE status NOT IN ('closed','rejected')) open_absence,(SELECT COUNT(*) FROM maintenance_tickets WHERE status NOT IN ('closed','verified')) open_tickets''').fetchone()
  zones=c.execute('''SELECT r.zone,COUNT(DISTINCT r.id) rooms,COALESCE(SUM(r.capacity),0) capacity,COUNT(w.id) workers FROM rooms r LEFT JOIN workers w ON w.room_no=r.room_no AND w.archived=0 WHERE r.usage_type='residential' GROUP BY r.zone ORDER BY CAST(r.zone AS INTEGER)''').fetchall()
  maint=c.execute('SELECT status,COUNT(*) count FROM maintenance_tickets GROUP BY status ORDER BY count DESC').fetchall();absence=c.execute('SELECT status,COUNT(*) count FROM absence_reports GROUP BY status ORDER BY count DESC').fetchall()
 occupancy=round((k['workers']/k['capacity']*100),1) if k['capacity'] else 0
 return page('''<h2>مركز التقارير ولوحة الإدارة</h2><div class="cards"><div class="card"><div>إجمالي العمال</div><div class="num">{{k.workers}}</div></div><div class="card"><div>الطاقة الاستيعابية</div><div class="num">{{k.capacity}}</div></div><div class="card"><div>نسبة الإشغال</div><div class="num">{{occupancy}}%</div></div><div class="card"><div>الغرف الفارغة</div><div class="num">{{k.vacant_rooms}}</div></div><div class="card"><div>بلاغات الغياب المفتوحة</div><div class="num">{{k.open_absence}}</div></div><div class="card"><div>بلاغات الصيانة المفتوحة</div><div class="num">{{k.open_tickets}}</div></div></div><div class="card"><h3>تقرير العمالة</h3><a class="btn" href="{{url_for('export_workers_excel')}}">Excel لجميع بيانات العمالة</a> <a class="btn btn2" href="{{url_for('export_workers_pdf')}}">PDF لجميع بيانات العمالة</a> <a class="btn btn2" href="{{url_for('workers')}}">فتح كشف العمال</a></div><div class="grid"><div class="card"><h3>الإشغال حسب الزون</h3><table class="tbl"><tr><th>الزون</th><th>الغرف</th><th>السعة</th><th>العمال</th><th>الإشغال</th></tr>{% for z in zones %}<tr><td>{{z.zone}}</td><td>{{z.rooms}}</td><td>{{z.capacity}}</td><td>{{z.workers}}</td><td>{{((z.workers/z.capacity*100)|round(1)) if z.capacity else 0}}%</td></tr>{% endfor %}</table></div><div class="card"><h3>حالات الصيانة</h3>{% for x in maint %}<p>{{status_ar.get(x.status,x.status)}}: <b>{{x.count}}</b></p>{% endfor %}<h3>حالات عدم التواجد</h3>{% for x in absence %}<p>{{x.status}}: <b>{{x.count}}</b></p>{% endfor %}</div></div>''','مركز التقارير',u,k=k,zones=zones,maint=maint,absence=absence,occupancy=occupancy,status_ar=STATUS_AR)

ABSENCE_REASONS=['سيتم إرساله إلى الحرم','مريض داخل السكن','تم نقله إلى المستشفى','غير متواجد في السكن','يوم راحة','إجازة سنوية','تم تغيير الوردية','رفض الذهاب إلى العمل','غياب بدون عذر','غير مقيم في السكن','منقول إلى مشروع آخر','خروج نهائي','أخرى']

def normalize_employee_no(value):
 if value is None:return ''
 text=str(value).strip().replace(' ','')
 if text.endswith('.0'):text=text[:-2]
 return ''.join(ch for ch in text if ch.isdigit()) or text

def read_absence_workbooks(files, valid_employee_nos=None):
 from openpyxl import load_workbook
 results={};errors=[]
 valid_employee_nos={normalize_employee_no(x) for x in (valid_employee_nos or []) if normalize_employee_no(x)}
 for f in files:
  if not f or not f.filename:continue
  if not f.filename.lower().endswith(('.xlsx','.xlsm')):
   errors.append(f'{f.filename}: الصيغة غير مدعومة');continue
  try:
   wb=load_workbook(f,read_only=True,data_only=True);found=0
   # لا نعتمد على اسم العمود أو مكانه. نفحص جميع خلايا جميع الأوراق،
   # ونقبل فقط الرقم الموجود فعلياً في قاعدة عمالة MAG CAMP أو عمالة المشروع.
   for ws in wb.worksheets:
    for row in ws.iter_rows(values_only=True):
     for raw in row:
      eno=normalize_employee_no(raw)
      if not eno or not eno.isdigit():continue
      if valid_employee_nos and eno not in valid_employee_nos:continue
      item=results.setdefault(eno,{'employee_no':eno,'file_names':[],'import_name':''})
      if f.filename not in item['file_names']:item['file_names'].append(f.filename)
      found+=1
   wb.close()
   if not found:errors.append(f'{f.filename}: لم يتم العثور على أرقام مطابقة لقاعدة بيانات العمالة')
  except Exception as exc:errors.append(f'{f.filename}: تعذر القراءة ({exc})')
 return list(results.values()),errors

@app.route('/attendance',methods=['GET','POST'])
@login_required
def attendance_batches():
 u=current_user()
 if not can_attendance(u):abort(403)
 err=''
 if request.method=='POST':
  files=request.files.getlist('absence_files')
  with closing(conn()) as lookup_db:
   valid_employee_nos={r['employee_no'] for r in lookup_db.execute("SELECT employee_no FROM workers WHERE COALESCE(archived,0)=0 AND employee_no IS NOT NULL UNION SELECT employee_no FROM project_workers WHERE employee_no IS NOT NULL").fetchall()}
  items,errors=read_absence_workbooks(files,valid_employee_nos)
  if not items:err='لم يتم العثور على أرقام وظيفية صالحة. '+'، '.join(errors)
  else:
   absence_date=request.form.get('absence_date') or date.today().isoformat();shift_label=request.form.get('shift_label','')
   with closing(conn()) as c:
    batch_no='ABS-'+datetime.now().strftime('%Y%m%d-%H%M%S')+'-'+uuid.uuid4().hex[:4].upper()
    cur=c.execute('INSERT INTO attendance_batches(batch_no,absence_date,shift_label,created_by,status,source_files_json,created_at) VALUES(?,?,?,?,?,?,?)',(batch_no,absence_date,shift_label,u['id'],'draft',json.dumps([f.filename for f in files if f and f.filename],ensure_ascii=False),now()));bid=cur.lastrowid
    for item in items:
     eno=item['employee_no'];resident=c.execute("SELECT employee_no,full_name,profession,zone,room_no FROM workers WHERE employee_no=? AND COALESCE(archived,0)=0 LIMIT 1",(eno,)).fetchone();project=c.execute('SELECT * FROM project_workers WHERE employee_no=?',(eno,)).fetchone();previous=c.execute("SELECT COUNT(*) FROM attendance_entries e JOIN attendance_batches b ON b.id=e.batch_id WHERE e.employee_no=? AND b.status='completed'",(eno,)).fetchone()[0]
     full_name=(resident['full_name'] if resident else None) or (project['full_name'] if project else None) or item.get('import_name') or 'غير معروف';profession=(resident['profession'] if resident else None) or (project['profession'] if project else None) or '';shift=(project['shift'] if project else None) or shift_label;rest_day=(project['rest_day'] if project else None) or ''
     c.execute('INSERT INTO attendance_entries(batch_id,employee_no,full_name,profession,shift,rest_day,is_resident,room_no,zone,source_files,previous_absences,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)',(bid,eno,full_name,profession,shift,rest_day,1 if resident else 0,resident['room_no'] if resident else '',str(resident['zone']) if resident else '','، '.join(item['file_names']),previous,now()))
    audit(c,u,'create','attendance_batch',bid,{'count':len(items),'files':len(files),'errors':errors});c.commit()
   return redirect(url_for('attendance_batch_detail',bid=bid))
 with closing(conn()) as c:
  if is_admin(u):rows=c.execute('SELECT b.*,u.display_name creator,(SELECT COUNT(*) FROM attendance_entries e WHERE e.batch_id=b.id) total FROM attendance_batches b LEFT JOIN users u ON u.id=b.created_by ORDER BY b.id DESC LIMIT 200').fetchall()
  else:rows=c.execute('SELECT b.*,u.display_name creator,(SELECT COUNT(*) FROM attendance_entries e WHERE e.batch_id=b.id) total FROM attendance_batches b LEFT JOIN users u ON u.id=b.created_by WHERE b.created_by=? ORDER BY b.id DESC LIMIT 100',(u['id'],)).fetchall()
 return page("""<h2>حصر غياب العمالة</h2><div class='card'><h3>إنشاء حصر جديد من عدة ملفات</h3>{% if err %}<p class='err'>{{err}}</p>{% endif %}<form method='post' enctype='multipart/form-data'><div class='grid'><div class='field'><label>تاريخ الغياب</label><input type='date' name='absence_date' value='{{today}}' required></div><div class='field'><label>الوردية العامة إن لم تكن مسجلة</label><input name='shift_label' placeholder='مثال: الوردية الأولى'></div></div><div class='field'><label>ملفات حصر الغياب — يمكن اختيار أكثر من Excel</label><input type='file' name='absence_files' accept='.xlsx,.xlsm' multiple required><small class='muted'>سيتم دمج جميع الملفات في حصر واحد، والرقم المكرر سيظهر مرة واحدة مع أسماء مصادره.</small></div><button class='btn'>رفع ودمج الملفات</button></form></div><div class='card'><h3>ملفات الحصر</h3><div class='tbl-wrap'><table class='tbl'><tr><th>رقم الحصر</th><th>التاريخ</th><th>الوردية</th><th>الإجمالي</th><th>أعده</th><th>الحالة</th><th></th></tr>{% for x in rows %}<tr><td>{{x.batch_no}}</td><td>{{x.absence_date}}</td><td>{{x.shift_label or '-'}}</td><td>{{x.total}}</td><td>{{x.creator or '-'}}</td><td><span class='badge'>{{'معتمد' if x.status=='completed' else 'مسودة'}}</span></td><td><a href='{{url_for("attendance_batch_detail",bid=x.id)}}'>فتح</a></td></tr>{% else %}<tr><td colspan='7'>لا توجد ملفات حصر حتى الآن.</td></tr>{% endfor %}</table></div></div>{% if admin %}<div class='card'><a class='btn btn2' href='{{url_for("attendance_history")}}'>سجل غياب العمالة</a> <a class='btn btn2' href='{{url_for("project_workers_import")}}'>تحديث ملف عمالة المشروع</a></div>{% endif %}""",'حصر الغياب',u,rows=rows,err=err,today=date.today().isoformat(),admin=is_admin(u))

@app.route('/attendance/<int:bid>',methods=['GET','POST'])
@login_required
def attendance_batch_detail(bid):
 u=current_user()
 if not can_attendance(u):abort(403)
 with closing(conn()) as c:
  b=c.execute('SELECT b.*,u.display_name creator FROM attendance_batches b LEFT JOIN users u ON u.id=b.created_by WHERE b.id=?',(bid,)).fetchone()
  if not b or (not is_admin(u) and b['created_by']!=u['id']):abort(404)
  if request.method=='POST':
   action=request.form.get('action')
   if action=='save' and b['status']!='completed':
    for eid in request.form.getlist('entry_id'):c.execute('UPDATE attendance_entries SET reason=?,notes=?,updated_at=? WHERE id=? AND batch_id=?',(request.form.get('reason_'+eid,''),request.form.get('notes_'+eid,''),now(),eid,bid))
    bulk=request.form.get('bulk_reason','');selected=request.form.getlist('selected')
    if bulk and selected:
     q=','.join('?'*len(selected));c.execute(f'UPDATE attendance_entries SET reason=?,updated_at=? WHERE batch_id=? AND id IN ({q})',[bulk,now(),bid]+selected)
    c.commit();return redirect(url_for('attendance_batch_detail',bid=bid))
   if action=='complete' and b['status']!='completed':
    missing=c.execute("SELECT COUNT(*) FROM attendance_entries WHERE batch_id=? AND COALESCE(reason,'')=''",(bid,)).fetchone()[0]
    if missing:return redirect(url_for('attendance_batch_detail',bid=bid,missing=missing))
    c.execute("UPDATE attendance_batches SET status='completed',completed_at=? WHERE id=?",(now(),bid));audit(c,u,'complete','attendance_batch',bid);c.commit();return redirect(url_for('attendance_batch_detail',bid=bid))
   if action=='reopen' and is_admin(u):c.execute("UPDATE attendance_batches SET status='draft',completed_at=NULL WHERE id=?",(bid,));c.commit();return redirect(url_for('attendance_batch_detail',bid=bid))
  rows=c.execute('SELECT * FROM attendance_entries WHERE batch_id=? ORDER BY previous_absences DESC,full_name',(bid,)).fetchall();summary=c.execute("SELECT COALESCE(NULLIF(reason,''),'غير مسجل') reason,COUNT(*) count FROM attendance_entries WHERE batch_id=? GROUP BY COALESCE(NULLIF(reason,''),'غير مسجل') ORDER BY count DESC",(bid,)).fetchall()
 missing=request.args.get('missing')
 return page("""<h2>حصر الغياب {{b.batch_no}}</h2><div class='cards'><div class='card'><div>إجمالي المتغيبين</div><div class='num'>{{rows|length}}</div></div><div class='card'><div>مقيمون في السكن</div><div class='num'>{{rows|selectattr('is_resident','equalto',1)|list|length}}</div></div><div class='card'><div>غير مقيمين</div><div class='num'>{{rows|selectattr('is_resident','equalto',0)|list|length}}</div></div><div class='card'><div>متكرر غيابهم</div><div class='num'>{{rows|selectattr('previous_absences','gt',0)|list|length}}</div></div></div>{% if missing %}<p class='err'>متبقي {{missing}} عامل بدون سبب غياب.</p>{% endif %}<div class='card'><p><b>التاريخ:</b> {{b.absence_date}} · <b>أعده:</b> {{b.creator}} · <b>الحالة:</b> {{'معتمد' if b.status=='completed' else 'مسودة'}}</p><a class='btn' href='{{url_for("attendance_export_excel",bid=b.id)}}'>Excel Dashboard</a> <a class='btn btn2' href='{{url_for("attendance_export_pdf",bid=b.id)}}'>PDF Dashboard</a>{% if admin and b.status=='completed' %}<form method='post' style='display:inline'><button class='btn danger' name='action' value='reopen'>إعادة فتح</button></form>{% endif %}</div><div class='card'><h3>ملخص الأسباب</h3>{% for x in summary %}<span class='badge'>{{x.reason}}: {{x.count}}</span> {% endfor %}</div><form method='post'><div class='card'><div class='grid'><div class='field'><label>تطبيق سبب على المحددين</label><select name='bulk_reason'><option value=''>اختر السبب</option>{% for r in reasons %}<option>{{r}}</option>{% endfor %}</select></div><div class='field'><label>&nbsp;</label><button class='btn' name='action' value='save'>حفظ التعديلات</button></div></div></div><div class='tbl-wrap'><table class='tbl'><tr><th>تحديد</th><th>الرقم</th><th>الاسم</th><th>الغرفة</th><th>الزون</th><th>الوردية</th><th>الراحة</th><th>السكن</th><th>الغياب السابق</th><th>مصدر الملف</th><th>السبب</th><th>ملاحظات</th></tr>{% for x in rows %}<tr><td><input type='checkbox' name='selected' value='{{x.id}}'><input type='hidden' name='entry_id' value='{{x.id}}'></td><td>{{x.employee_no}}</td><td>{{x.full_name}}</td><td>{{x.room_no or '-'}}</td><td>{{x.zone or '-'}}</td><td>{{x.shift or '-'}}</td><td>{{x.rest_day or '-'}}</td><td>{{'مقيم' if x.is_resident else 'غير مقيم'}}</td><td><b>{{x.previous_absences}}</b>{% if x.previous_absences>=5 %}<br><span class='badge red'>غياب متكرر</span>{% endif %}</td><td>{{x.source_files}}</td><td><select name='reason_{{x.id}}' {% if b.status=='completed' %}disabled{% endif %}><option value=''>اختر</option>{% for r in reasons %}<option {% if x.reason==r %}selected{% endif %}>{{r}}</option>{% endfor %}</select></td><td><input name='notes_{{x.id}}' value='{{x.notes or ""}}' {% if b.status=='completed' %}disabled{% endif %}></td></tr>{% endfor %}</table></div>{% if b.status!='completed' %}<div class='card'><button class='btn' name='action' value='save'>حفظ</button> <button class='btn btn2' name='action' value='complete'>اعتماد الحصر النهائي</button></div>{% endif %}</form>""",'تفاصيل حصر الغياب',u,b=b,rows=rows,summary=summary,reasons=ABSENCE_REASONS,missing=missing,admin=is_admin(u))

@app.route('/attendance/project-workers',methods=['GET','POST'])
@login_required
def project_workers_import():
 u=current_user()
 if not is_admin(u):abort(403)
 msg='';err=''
 if request.method=='POST':
  f=request.files.get('project_file')
  try:
   from openpyxl import load_workbook
   wb=load_workbook(f,read_only=True,data_only=True);count=0
   with closing(conn()) as c:
    for ws in wb.worksheets:
     header=None
     for idx,row in enumerate(ws.iter_rows(min_row=1,max_row=min(ws.max_row,15),values_only=True),1):
      vals=[str(v or '').strip() for v in row]
      if any(v in ('الرقم الوظيفي','رقم وظيفي') for v in vals):header=(idx,vals);break
     if not header:continue
     hi,vals=header
     def col(names):
      for n in names:
       if n in vals:return vals.index(n)
      return None
     ci=col(['الرقم الوظيفي','رقم وظيفي']);cn=col(['الإسم','الاسم']);cp=col(['المسمى الوظيفي']);cs=col(['الوردية']);cr=col(['الراحات','الراحة']);cl=col(['الموقع','الموقع الاساسي '])
     for row in ws.iter_rows(min_row=hi+1,values_only=True):
      eno=normalize_employee_no(row[ci] if ci is not None and len(row)>ci else None)
      if not eno:continue
      val=lambda i: str(row[i]).strip() if i is not None and len(row)>i and row[i] is not None else ''
      c.execute('INSERT INTO project_workers(employee_no,full_name,profession,shift,rest_day,source_sheet,source_location,updated_at) VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(employee_no) DO UPDATE SET full_name=excluded.full_name,profession=excluded.profession,shift=excluded.shift,rest_day=excluded.rest_day,source_sheet=excluded.source_sheet,source_location=excluded.source_location,updated_at=excluded.updated_at',(eno,val(cn),val(cp),val(cs),val(cr),ws.title,val(cl),now()));count+=1
    audit(c,u,'import','project_workers',None,{'count':count,'file':f.filename});c.commit()
   wb.close();msg=f'تم دمج وتحديث {count} سجل من عمالة المشروع.'
  except Exception as exc:err='تعذر استيراد الملف: '+str(exc)
 with closing(conn()) as c:total=c.execute('SELECT COUNT(*) FROM project_workers').fetchone()[0];matched=c.execute('SELECT COUNT(*) FROM project_workers p WHERE EXISTS(SELECT 1 FROM workers w WHERE w.employee_no=p.employee_no AND w.archived=0)').fetchone()[0]
 return page("""<h2>تحديث عمالة المشروع والورديات والراحات</h2>{% if msg %}<p class='ok'>{{msg}}</p>{% endif %}{% if err %}<p class='err'>{{err}}</p>{% endif %}<div class='cards'><div class='card'><div>عمالة المشروع</div><div class='num'>{{total}}</div></div><div class='card'><div>مطابقون مع سكان السكن</div><div class='num'>{{matched}}</div></div><div class='card'><div>غير مقيمين في السكن</div><div class='num'>{{total-matched}}</div></div></div><div class='card'><form method='post' enctype='multipart/form-data'><div class='field'><label>ملف جميع عمالة المشروع</label><input type='file' name='project_file' accept='.xlsx,.xlsm' required></div><button class='btn'>دمج وتحديث البيانات</button></form></div>""",'عمالة المشروع',u,total=total,matched=matched,msg=msg,err=err)

@app.get('/attendance/history')
@login_required
def attendance_history():
 u=current_user()
 if not is_admin(u):abort(403)
 q=request.args.get('q','').strip();params=[];where="b.status='completed'"
 if q:where+=' AND (e.employee_no LIKE ? OR e.full_name LIKE ?)';params=[f'%{q}%',f'%{q}%']
 with closing(conn()) as c:rows=c.execute(f'SELECT e.employee_no,MAX(e.full_name) full_name,MAX(e.room_no) room_no,MAX(e.shift) shift,COUNT(*) total,MAX(b.absence_date) last_date,GROUP_CONCAT(DISTINCT e.reason) reasons FROM attendance_entries e JOIN attendance_batches b ON b.id=e.batch_id WHERE {where} GROUP BY e.employee_no ORDER BY total DESC,last_date DESC LIMIT 1000',params).fetchall()
 return page("""<h2>سجل غياب العمالة</h2><form class='card'><div class='field'><label>بحث بالرقم الوظيفي أو الاسم</label><input name='q' value='{{q}}'></div><button class='btn'>بحث</button></form><div class='tbl-wrap'><table class='tbl'><tr><th>الرقم</th><th>الاسم</th><th>الغرفة</th><th>الوردية</th><th>إجمالي الغياب</th><th>آخر غياب</th><th>الأسباب المسجلة</th></tr>{% for x in rows %}<tr><td>{{x.employee_no}}</td><td>{{x.full_name}}</td><td>{{x.room_no or '-'}}</td><td>{{x.shift or '-'}}</td><td><b>{{x.total}}</b>{% if x.total>=5 %} <span class='badge red'>متكرر</span>{% endif %}</td><td>{{x.last_date}}</td><td>{{x.reasons or '-'}}</td></tr>{% endfor %}</table></div>""",'سجل الغياب',u,rows=rows,q=q)

def attendance_export_data(bid,u):
 with closing(conn()) as c:
  b=c.execute('SELECT b.*,u.display_name creator FROM attendance_batches b LEFT JOIN users u ON u.id=b.created_by WHERE b.id=?',(bid,)).fetchone()
  if not b or (not is_admin(u) and b['created_by']!=u['id']):abort(404)
  rows=c.execute('SELECT * FROM attendance_entries WHERE batch_id=? ORDER BY previous_absences DESC,full_name',(bid,)).fetchall();summary=c.execute("SELECT COALESCE(NULLIF(reason,''),'غير مسجل') reason,COUNT(*) count FROM attendance_entries WHERE batch_id=? GROUP BY COALESCE(NULLIF(reason,''),'غير مسجل') ORDER BY count DESC",(bid,)).fetchall()
 return b,rows,summary

@app.get('/attendance/<int:bid>/export.xlsx')
@login_required
def attendance_export_excel(bid):
 u=current_user()
 if not can_attendance(u):abort(403)
 from openpyxl import Workbook
 from openpyxl.styles import Font,PatternFill,Alignment
 from openpyxl.chart import PieChart,Reference
 b,rows,summary=attendance_export_data(bid,u);wb=Workbook();dash=wb.active;dash.title='Dashboard';details=wb.create_sheet('تفاصيل الغياب');repeat=wb.create_sheet('متكرر الغياب')
 for ws in (dash,details,repeat):ws.sheet_view.rightToLeft=True
 dash.merge_cells('A1:H2');dash['A1']='MAG CAMP — لوحة حصر غياب العمالة';dash['A1'].font=Font(size=18,bold=True,color='FFFFFF');dash['A1'].fill=PatternFill('solid',fgColor='17365D');dash['A1'].alignment=Alignment(horizontal='center',vertical='center')
 resident=sum(1 for x in rows if x['is_resident']);repeated=sum(1 for x in rows if x['previous_absences']>0);kpis=[('إجمالي المتغيبين',len(rows)),('المقيمون',resident),('غير المقيمين',len(rows)-resident),('لديهم غياب سابق',repeated)]
 for i,(label,value) in enumerate(kpis):
  col=1+i*2;dash.cell(4,col,label);dash.cell(5,col,value);dash.merge_cells(start_row=4,start_column=col,end_row=4,end_column=col+1);dash.merge_cells(start_row=5,start_column=col,end_row=5,end_column=col+1)
  for r in (4,5):dash.cell(r,col).alignment=Alignment(horizontal='center');dash.cell(r,col).fill=PatternFill('solid',fgColor='D9EAF7' if r==4 else 'FFFFFF');dash.cell(r,col).font=Font(bold=True,size=14 if r==5 else 11)
 dash['A7']='السبب';dash['B7']='العدد'
 for i,x in enumerate(summary,8):dash.cell(i,1,x['reason']);dash.cell(i,2,x['count'])
 if summary:
  pie=PieChart();pie.title='توزيع أسباب الغياب';pie.add_data(Reference(dash,min_col=2,min_row=7,max_row=7+len(summary)),titles_from_data=True);pie.set_categories(Reference(dash,min_col=1,min_row=8,max_row=7+len(summary)));pie.height=8;pie.width=12;dash.add_chart(pie,'D7')
 headers=['م','الرقم الوظيفي','الاسم','المهنة','الغرفة','الزون','الوردية','الراحة','حالة السكن','الغياب السابق','مصدر الملف','سبب الغياب','الملاحظات'];details.append(headers);repeat.append(headers)
 for i,x in enumerate(rows,1):
  line=[i,x['employee_no'],x['full_name'],x['profession'],x['room_no'],x['zone'],x['shift'],x['rest_day'],'مقيم' if x['is_resident'] else 'غير مقيم',x['previous_absences'],x['source_files'],x['reason'],x['notes']];details.append(line)
  if x['previous_absences']>0:repeat.append(line)
 for ws in (details,repeat):
  for cell in ws[1]:cell.fill=PatternFill('solid',fgColor='17365D');cell.font=Font(color='FFFFFF',bold=True);cell.alignment=Alignment(horizontal='center')
  ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions
  for idx,w in enumerate([6,16,28,22,12,10,16,14,14,14,30,24,30],1):ws.column_dimensions[chr(64+idx)].width=w
 for col in range(1,9):dash.column_dimensions[chr(64+col)].width=18
 bio=io.BytesIO();wb.save(bio);bio.seek(0)
 return send_file(bio,as_attachment=True,download_name=f'attendance_{b["absence_date"]}_{b["batch_no"]}.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get('/attendance/<int:bid>/export.pdf')
@login_required
def attendance_export_pdf(bid):
 u=current_user()
 if not can_attendance(u):abort(403)
 from reportlab.lib.pagesizes import A3,landscape
 from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer,PageBreak
 from reportlab.lib import colors
 from reportlab.lib.styles import ParagraphStyle
 from reportlab.pdfbase import pdfmetrics
 from reportlab.pdfbase.ttfonts import TTFont
 b,rows,summary=attendance_export_data(bid,u);bio=io.BytesIO()
 try:pdfmetrics.registerFont(TTFont('Arabic','/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
 except Exception:pass
 def ar(v):return ar_text(v)
 doc=SimpleDocTemplate(bio,pagesize=landscape(A3),rightMargin=24,leftMargin=24,topMargin=24,bottomMargin=24);style=ParagraphStyle('ar',fontName='Arabic',fontSize=10,alignment=2);title=ParagraphStyle('title',fontName='Arabic',fontSize=18,alignment=1,spaceAfter=16)
 story=[Paragraph(ar('MAG CAMP — تقرير حصر غياب العمالة'),title),Paragraph(ar(f'رقم الحصر: {b["batch_no"]} | التاريخ: {b["absence_date"]} | أعده: {b["creator"]}'),style),Spacer(1,12)]
 kpi=[[ar('إجمالي المتغيبين'),ar('المقيمون'),ar('غير المقيمين'),ar('لديهم غياب سابق')],[len(rows),sum(1 for x in rows if x['is_resident']),sum(1 for x in rows if not x['is_resident']),sum(1 for x in rows if x['previous_absences']>0)]];t=Table(kpi,colWidths=[180]*4,rowHeights=[30,38]);t.setStyle(TableStyle([('FONT',(0,0),(-1,-1),'Arabic'),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#17365D')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('ALIGN',(0,0),(-1,-1),'CENTER'),('GRID',(0,0),(-1,-1),0.5,colors.grey)]));story += [t,Spacer(1,16)]
 sm=[[ar('سبب الغياب'),ar('العدد')]]+[[ar(x['reason']),x['count']] for x in summary];st=Table(sm,colWidths=[350,100]);st.setStyle(TableStyle([('FONT',(0,0),(-1,-1),'Arabic'),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#17365D')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('ALIGN',(0,0),(-1,-1),'CENTER'),('GRID',(0,0),(-1,-1),0.4,colors.grey)]));story += [st,PageBreak()]
 data=[[ar(x) for x in ['م','الرقم','الاسم','الغرفة','الزون','الوردية','الراحة','السكن','السابق','السبب']]]
 for i,x in enumerate(rows,1):data.append([i,x['employee_no'],ar(x['full_name']),x['room_no'] or '-',x['zone'] or '-',ar(x['shift']),ar(x['rest_day']),ar('مقيم' if x['is_resident'] else 'غير مقيم'),x['previous_absences'],ar(x['reason'])])
 dt=Table(data,repeatRows=1,colWidths=[30,70,180,55,45,80,70,70,45,140]);dt.setStyle(TableStyle([('FONT',(0,0),(-1,-1),'Arabic'),('FONTSIZE',(0,0),(-1,-1),7),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#17365D')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.25,colors.grey),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(-1,-1),'CENTER')]));story.append(dt);doc.build(story);bio.seek(0)
 return send_file(bio,as_attachment=True,download_name=f'attendance_{b["absence_date"]}_{b["batch_no"]}.pdf',mimetype='application/pdf')


@app.route('/backup',methods=['GET','POST'])
@login_required
def backup_restore():
 u=current_user()
 if not is_admin(u):abort(403)
 msg='';err=''
 if request.method=='POST':
  f=request.files.get('database_file')
  if not f or not f.filename.lower().endswith('.db'):err='اختر ملف قاعدة بيانات بصيغة DB.'
  else:
   fd,tmp=tempfile.mkstemp(suffix='.db');os.close(fd);f.save(tmp)
   try:
    test=sqlite3.connect(tmp);integrity=test.execute('PRAGMA integrity_check').fetchone()[0];required={'users','rooms','workers'};tables={x[0] for x in test.execute("SELECT name FROM sqlite_master WHERE type='table'")};test.close()
    if integrity!='ok' or not required.issubset(tables):raise ValueError('ملف قاعدة البيانات غير صالح أو لا يحتوي الجداول الأساسية.')
    backup_name=DB+'.before_restore_'+datetime.now().strftime('%Y%m%d_%H%M%S');shutil.copy2(DB,backup_name);shutil.copy2(tmp,DB);ensure_schema();msg='تمت استعادة قاعدة البيانات بنجاح، مع إنشاء نسخة احتياطية تلقائية قبل الاستعادة.'
   except Exception as e:err='تعذر الاستعادة: '+str(e)
   finally:
    try:os.remove(tmp)
    except OSError:pass
 return page('''<div class="card"><h2>النسخ الاحتياطي والاستعادة</h2>{% if msg %}<p class="ok">{{msg}}</p>{% endif %}{% if err %}<p class="err">{{err}}</p>{% endif %}<p><a class="btn" href="{{url_for('download_backup')}}">تنزيل نسخة احتياطية الآن</a></p><hr><h3>استعادة قاعدة البيانات</h3><p class="muted">سيتم إنشاء نسخة تلقائية من القاعدة الحالية قبل الاستعادة. استخدم ملف DB صادر من نفس النظام.</p><form method="post" enctype="multipart/form-data"><input type="file" name="database_file" accept=".db" required><button class="btn danger">استعادة</button></form></div>''','النسخ الاحتياطي',u,msg=msg,err=err)

@app.get('/backup/download')
@login_required
def download_backup():
 u=current_user()
 if not is_admin(u):abort(403)
 fd,tmp=tempfile.mkstemp(suffix='.db');os.close(fd)
 with closing(conn()) as source:
  target=sqlite3.connect(tmp);source.backup(target);target.close()
 with closing(conn()) as c:audit(c,u,'backup','database',None);c.commit()
 return send_file(tmp,as_attachment=True,download_name=f'mag_camp_backup_{datetime.now().strftime("%Y%m%d_%H%M")}.db',mimetype='application/octet-stream',max_age=0)

@app.get('/notifications')
@login_required
def notifications():
 u=current_user();items=[]
 with closing(conn()) as c:
  if is_admin(u):
   for x in c.execute("SELECT id,action_no,created_at FROM housing_actions WHERE final_status NOT IN ('approved','rejected') ORDER BY id DESC LIMIT 30").fetchall():items.append({'title':'طلب سكن بانتظار الاعتماد','text':x['action_no'],'at':x['created_at'],'url':url_for('worker_change_request_detail',qid=x['id'])})
  ticket_where="status NOT IN ('closed','verified')";params=[]
  if u['role'] in ('housing_supervisor','housing_monitor'):ticket_where+=' AND reported_by=?';params.append(u['id'])
  for x in c.execute(f"SELECT id,ticket_no,status,created_at FROM maintenance_tickets WHERE {ticket_where} ORDER BY id DESC LIMIT 30",params).fetchall():items.append({'title':'بلاغ صيانة مفتوح','text':x['ticket_no']+' · '+STATUS_AR.get(x['status'],x['status']),'at':x['created_at'],'url':url_for('ticket_detail',tid=x['id'])})
 items=sorted(items,key=lambda x:x['at'] or '',reverse=True)
 return page('''<h2>{{t.notifications}}</h2>{% for x in items %}<a class="card" style="display:block;color:inherit;text-decoration:none" href="{{x.url}}"><b>{{x.title}}</b><p>{{x.text}}</p><small>{{x.at}}</small></a>{% else %}<div class="card">لا توجد إشعارات حالية.</div>{% endfor %}''',tr('notifications'),u,items=items)

@app.get('/users')
@login_required
def users():
 u=current_user()
 if not is_admin(u):abort(403)
 with closing(conn()) as c:rows=c.execute("SELECT u.*,GROUP_CONCAT(a.room_text,'، ') room_text FROM users u LEFT JOIN assignments a ON a.user_id=u.id GROUP BY u.id ORDER BY u.active DESC,u.role,u.display_name").fetchall()
 return page('''<h2>المستخدمون والصلاحيات</h2><table class="tbl"><tr><th>الرقم</th><th>الاسم</th><th>الدور</th><th>النطاق</th><th>الحالة</th></tr>{% for x in rows %}<tr><td>{{x.employee_no}}</td><td>{{x.display_name}}</td><td>{{roles.get(x.role,x.role)}}</td><td>{{x.room_text or '-'}}</td><td>{{'نشط' if x.active else 'موقوف'}}</td></tr>{% endfor %}</table>''','المستخدمون',u,rows=rows,roles=ROLE_AR)
@app.get('/audit-logs')
@login_required
def audit_logs():
 u=current_user()
 if not is_admin(u):abort(403)
 with closing(conn()) as c:rows=c.execute('SELECT * FROM audit_logs ORDER BY id DESC LIMIT 500').fetchall()
 return page('''<h2>سجل العمليات</h2><table class="tbl"><tr><th>التاريخ</th><th>المستخدم</th><th>العملية</th><th>النوع</th><th>الرقم</th><th>التفاصيل</th></tr>{% for x in rows %}<tr><td>{{x.created_at}}</td><td>{{x.username}}</td><td>{{x.action}}</td><td>{{x.entity_type}}</td><td>{{x.entity_id}}</td><td>{{x.details_json}}</td></tr>{% endfor %}</table>''','سجل العمليات',u,rows=rows)
@app.errorhandler(RequestEntityTooLarge)
def upload_too_large(e):
 u=current_user();limit=int(app.config['MAX_CONTENT_LENGTH']/1024/1024)
 return page('<div class="card"><h2>حجم الصور كبير</h2><p class="err">تجاوزت المرفقات الحد المسموح وهو {{limit}} MB للطلب الواحد.</p><p>قلل عدد الصور أو حجمها ثم أعد المحاولة.</p></div>','خطأ رفع الصور',u,limit=limit),413
@app.errorhandler(500)
def internal_error(e):
 return page('<div class="card"><h2>تعذر تنفيذ العملية</h2><p class="err">حدث خطأ غير متوقع. لم يتم اعتماد العملية، ويمكنك إعادة المحاولة.</p></div>','خطأ في النظام',current_user()),500
@app.errorhandler(403)
def forbidden(e):return page('<div class="card"><h2>غير مصرح</h2><p>ليس لديك صلاحية لفتح هذه الصفحة.</p></div>','غير مصرح',current_user()),403
@app.errorhandler(404)
def notfound(e):return page('<div class="card"><h2>غير موجود</h2></div>','غير موجود',current_user()),404
if __name__=='__main__':app.run(host='0.0.0.0',port=int(os.environ.get('PORT','10000')))
